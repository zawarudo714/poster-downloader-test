"""
Admin-facing routes. Mounted under /admin.

Endpoints:
    GET  /admin                           → dashboard (counts + per-user stats)
    GET  /admin/users                     → user management page
    POST /admin/users/create              → create user
    POST /admin/users/{id}/toggle         → toggle is_active
    POST /admin/users/{id}/reset_password → set new password
    POST /admin/users/{id}/release_queue  → release ALL claims (started or not) back to pending
    GET  /admin/master                    → paginated master sheet view
    GET  /admin/api/master                → JSON paginated master (same paging used by user/master)
    POST /admin/master/{id}/status        → set status for one row (admin override)
    POST /admin/master/bulk_status        → bulk status change for filtered/selected ids
    POST /admin/master/clear              → wipe all master rows
    POST /admin/master/upload             → start background import
    GET  /admin/import/{id}               → poll import status
    GET  /admin/browse                    → image browser shell
    GET  /admin/api/browse                → JSON of all live posters for (worker, date)
    GET  /admin/file/{poster_id}          → serve a SavedPoster image
    POST /admin/poster/{id}/flag          → flag a SavedPoster
    POST /admin/poster/{id}/unflag        → resolve all open flags on a poster
    GET  /admin/revisions                 → revisions list page
    GET  /admin/audit                     → activity log page
    GET  /admin/api/audit                 → activity log JSON
    GET  /admin/api/tree                  → workspace fs tree
    POST /admin/zip/start                 → start background zip job
    GET  /admin/zip/status/{job_id}       → poll zip job
    GET  /admin/zip/download/{job_id}     → download finished zip
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import threading
import time
import zipfile
from datetime import date, date as date_type, datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session

from ..audit import log as log_activity
from ..auth import hash_password, require_admin
from ..config import MASTER_PAGE_SIZE
from ..config import WORKSPACE_DIR
from ..db import SessionLocal, get_db
from ..models import (
    ActivityLog, AppSetting, ChatMessage, ChatReadState,
    ImportJob, MasterTitle, PaymentRun, ProcessedImage, Project, Revision,
    SavedPoster, UploadAccount, UploadTracking, User, UserProject,
)
from ..pipeline import ensure_default_project
from ..projects import (
    active_project, allowed_projects, project_by_slug, remember_project,
    scope_titles, set_project_cookie,
)
from ..timeutil import fmt_local, local_today
from ..templating import templates
from ..utils import (
    count_live_posters_for_master,
    count_user_saves_for_date, count_user_saves_for_week,
    list_date_folders, list_users_with_workspaces, safe_under_workspace,
    saved_poster_folder, saved_poster_path,
)


router = APIRouter(prefix="/admin")


# ── Project switching ────────────────────────────────────────────────────────
#
# The active project is session state, not part of the URL — see the module
# docstring in app/projects.py for why. These two routes are the only way it
# changes, which keeps the "where am I" logic in one place.

@router.get("/project/{slug}")
def enter_project(
    slug: str,
    request: Request,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Enter a project. The nav replaces itself from here on."""
    proj = project_by_slug(db, slug)
    if proj is None:
        raise HTTPException(status_code=404, detail="No such project")

    resp = RedirectResponse(url="/admin/browse", status_code=303)
    set_project_cookie(resp, proj)
    remember_project(db, admin, proj)
    db.commit()
    return resp


@router.get("/projects/exit")
def exit_project(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Back out to the master level."""
    resp = RedirectResponse(url="/admin", status_code=303)
    set_project_cookie(resp, None)
    remember_project(db, admin, None)
    db.commit()
    return resp


def current_project(request: Request, admin: User, db: Session) -> Project:
    """
    The project a project-scoped route operates on.

    Falls back to the default project rather than erroring, because every
    existing page was written before projects existed and its behaviour on a
    single-project install must be byte-identical to what it does today.
    """
    return active_project(request, db, admin) or ensure_default_project(db)


# ── Diagnostics ──────────────────────────────────────────────────────────────

@router.get("/diagnostics", response_class=HTMLResponse)
def diagnostics_page(request: Request, admin: User = Depends(require_admin),
                     db: Session = Depends(get_db)):
    """
    Master-level: spans every project by default.

    That is the useful default — the question this page answers is "is
    anything wrong anywhere", and having to visit each niche in turn to ask
    it would guarantee the quiet one goes unchecked. The filter narrows it
    when you already know where you are looking.
    """
    return templates.TemplateResponse(
        request, "admin_diagnostics.html",
        {"user": admin, "admin": admin, "active_tab": "diagnostics",
         "projects": db.query(Project).order_by(Project.id.asc()).all()},
    )


@router.get("/api/diagnostics")
def api_diagnostics(project_id: int = Query(0),
                    admin: User = Depends(require_admin),
                    db: Session = Depends(get_db)):
    """
    Run the consistency scan, optionally for one project only.

    Fetched on demand rather than on page load: the disk walk is the slow part
    and it should be something the admin chooses to start, not something that
    fires every time they land on the page.
    """
    from ..diagnostics import run_all
    return JSONResponse(run_all(db, project_id=project_id or None))


# ── Dashboard ────────────────────────────────────────────────────────────────

@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
def admin_dashboard(request: Request, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    today = local_today()
    users = db.query(User).filter_by(role="worker", is_deleted=0).order_by(User.username.asc()).all()
    user_stats = []
    for w in users:
        claims = (
            db.query(MasterTitle)
              .filter(MasterTitle.claimed_by_id == w.id, MasterTitle.status == "in_progress")
              .count()
        )
        user_stats.append({
            "id": w.id,
            "username": w.username,
            "today":  count_user_saves_for_date(db, w.username, today),
            "week":   count_user_saves_for_week(db, w.username, today),
            "claims": claims,
            "is_active": bool(w.is_active),
            "presence": _presence_label(w.last_seen_at),
        })
    open_revs = db.query(Revision).filter_by(status="open").count()
    awaiting_revs = db.query(Revision).filter_by(status="awaiting_approval").count()

    # Master totals
    pending_master   = db.query(MasterTitle).filter_by(status="pending").count()
    in_progress      = db.query(MasterTitle).filter_by(status="in_progress").count()
    complete_master  = db.query(MasterTitle).filter_by(status="complete").count()
    complete_pending = db.query(MasterTitle).filter_by(status="complete_pending").count()
    skipped_master   = db.query(MasterTitle).filter_by(status="skipped").count()
    revision_master  = db.query(MasterTitle).filter(MasterTitle.needs_revision == 1).count()
    total_master     = db.query(MasterTitle).count()

    # ── "Needs your attention" digest counts ────────────────────────────────
    # Deletion review = revision auto-resolved by worker delete that admin
    # hasn't yet acknowledged or escalated. Same logic as the revisions page.
    pending_deletions = (
        db.query(Revision)
          .filter(
              Revision.status == "resolved",
              Revision.admin_verdict.like("auto-resolved: %file deleted%"),
              Revision.admin_acked_at.is_(None),
          )
          .count()
    )

    # Unread chats for this admin viewer.
    from ..chat import admin_thread_summaries
    threads = admin_thread_summaries(db, viewer_id=admin.id)
    unread_chats = sum(t["unread"] for t in threads)

    # Owed since last payment per worker — sum eligible posters across
    # all workers using current rate (matches what /admin/payments would show
    # if you hit "this week" preset). Cap query: we only care about a count.
    from ..payments import eligible_poster_ids, get_rate_kes, parse_decimal, week_bounds_containing, get_week_start_day
    week_start_day = get_week_start_day(db)
    week_start, _ = week_bounds_containing(today, week_start_day)
    rate_dec = parse_decimal(get_rate_kes(db))
    eligible_total = 0
    workers_with_eligible = 0
    for w in users:
        n = len(eligible_poster_ids(db, worker_id=w.id, start=week_start, end=today))
        eligible_total += n
        if n > 0:
            workers_with_eligible += 1
    eligible_amount = str(rate_dec * eligible_total)
    if "." in eligible_amount:
        eligible_amount = eligible_amount.rstrip("0").rstrip(".") or "0"

    digest = {
        "awaiting_approval": awaiting_revs,
        "pending_completions": complete_pending,
        "pending_deletions": pending_deletions,
        "unread_chats":      unread_chats,
        "owed_amount_kes":   eligible_amount,
        "owed_count":        eligible_total,
        "owed_worker_count": workers_with_eligible,
        "week_start":        week_start.isoformat(),
        "week_end":          today.isoformat(),
    }

    # ── Per-project cards ───────────────────────────────────────────────────
    # The master dashboard's job is to answer "which project needs me?", so
    # each card carries the numbers you'd otherwise have to enter the project
    # to see. Counts are grouped in two queries rather than one per project —
    # with ten pipelines the naive version is 40 queries per page load.
    projects = allowed_projects(db, admin)
    status_counts: dict[tuple[int | None, str], int] = {
        (pid, st): n
        for pid, st, n in db.query(
            MasterTitle.project_id, MasterTitle.status, func.count(MasterTitle.id)
        ).group_by(MasterTitle.project_id, MasterTitle.status).all()
    }
    default_proj_id = ensure_default_project(db).id

    def _count(proj_id: int, status: str) -> int:
        # NULL project_id means the default project — the 101k imported rows
        # have never been backfilled, and treating NULL as "unassigned" here
        # would show the primary project as empty.
        n = status_counts.get((proj_id, status), 0)
        if proj_id == default_proj_id:
            n += status_counts.get((None, status), 0)
        return n

    project_cards = [
        {
            "id":       p.id,
            "slug":     p.slug,
            "name":     p.name,
            "site":     p.target_site,
            "source":   p.source_site,
            "pending":  _count(p.id, "pending"),
            "active":   _count(p.id, "in_progress"),
            "awaiting": _count(p.id, "complete_pending"),
            "complete": _count(p.id, "complete"),
            "skipped":  _count(p.id, "skipped"),
        }
        for p in projects
    ]

    # ── Shared machines, checked at MASTER level ─────────────────────────
    #
    # A worker node is not per-project. One Windows box does the Photoshop
    # work for the movie project AND the marketplace uploads for every
    # project, MUSIK included. So "is there a machine running" is an
    # account-wide question and belongs on the first page you land on, not
    # inside one project's tab where the other projects would never see it.
    #
    # This was originally written as a per-project check gated on
    # processor == 'photoshop'. Standing in MUSIK, whose processor is 'gpt',
    # it therefore stayed silent while the very node MUSIK uploads through
    # was dead.
    from ..models import WorkerNode

    node_stale_after = datetime.utcnow() - timedelta(minutes=5)
    all_nodes = db.query(WorkerNode).filter(WorkerNode.is_enabled == 1).all()
    offline_nodes = [
        {
            "name": n.name,
            "last_seen": fmt_local(n.last_seen_at, "%Y-%m-%d %H:%M") if n.last_seen_at else None,
            "age_s": (int((datetime.utcnow() - n.last_seen_at).total_seconds())
                      if n.last_seen_at else None),
        }
        for n in all_nodes
        if not (n.last_seen_at and n.last_seen_at > node_stale_after)
    ]

    # What is actually held up by it, so the banner states a consequence
    # rather than just a status.
    node_waiting = {
        "processing": (
            db.query(func.count(SavedPoster.id))
              .filter(SavedPoster.pipeline_status == "greenlit",
                      SavedPoster.deleted_at.is_(None))
              .scalar() or 0
        ),
        "uploads": (
            db.query(func.count(UploadTracking.id))
              .filter(UploadTracking.status.in_(("pending", "failed")))
              .scalar() or 0
        ),
    }

    return templates.TemplateResponse(
        request,
        "admin_dashboard.html",
        {"user": admin, "admin": admin, "today": today.isoformat(),
            "projects": project_cards,
            "node_total": len(all_nodes),
            "offline_nodes": offline_nodes,
            "node_waiting": node_waiting,
            "users": user_stats, "open_revisions": open_revs,
            "awaiting_revisions": awaiting_revs,
            "master_pending": pending_master, "master_in_progress": in_progress,
            "master_complete": complete_master,
            "master_complete_pending": complete_pending,
            "master_skipped": skipped_master,
            "master_needs_revision": revision_master, "master_total": total_master,
            "digest": digest,
            "active_tab": "dashboard",
        },
    )


def _presence_label(last_seen_at) -> str:
    """Convert last_seen_at (UTC datetime) → 'online' | 'away' | 'offline' | 'never'."""
    if last_seen_at is None:
        return "never"
    delta = datetime.utcnow() - last_seen_at
    secs = delta.total_seconds()
    if secs < 60:    return "online"
    if secs < 600:   return "away"        # 1–10 min
    return "offline"


# ── User management ─────────────────────────────────────────────────────────

@router.get("/users", response_class=HTMLResponse)
def users_page(request: Request, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    """
    Users live in the live users table — single env, simpler than before.
    Claim counts come straight from the same DB. Soft-deleted users
    (is_deleted=1) are hidden by default.
    """
    users = (
        db.query(User)
          .filter(User.is_deleted == 0)
          .order_by(User.role.desc(), User.username.asc())
          .all()
    )
    claim_counts = dict(
        db.query(MasterTitle.claimed_by_id, func.count(MasterTitle.id))
          .filter(MasterTitle.claimed_by_id.isnot(None),
                  MasterTitle.status == "in_progress")
          .group_by(MasterTitle.claimed_by_id)
          .all()
    )

    # Project assignments, one query for everyone rather than one per user.
    all_projects = db.query(Project).filter(Project.is_active == 1).order_by(Project.id).all()
    assigned: dict[int, set[int]] = {}
    for row in db.query(UserProject).all():
        assigned.setdefault(row.user_id, set()).add(row.project_id)
    return templates.TemplateResponse(
        request,
        "admin_users.html",
        {"user": admin, "admin": admin, "users": users,
         "claim_counts": claim_counts,
         "all_projects": all_projects,
         "assigned": {uid: sorted(pids) for uid, pids in assigned.items()},
         "active_tab": "users"},
    )


@router.post("/users/{user_id}/projects")
def set_user_projects(
    user_id: int,
    project_ids: str = Form(""),   # comma-separated; empty = every project
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Set which projects a worker draws work from.

    An EMPTY list deliberately means "no restriction" rather than "no work".
    Every worker in the database today has no rows here, and the upgrade must
    not silently stop them working — so absence of a rule means absence of a
    restriction, and turning a worker off is what the disable button is for.

    Claims already held are left alone. A worker mid-way through titles from
    a project they're being removed from should finish them; yanking the work
    out from under someone is how posters end up half-saved with nobody
    responsible for them.
    """
    u = db.query(User).filter_by(id=user_id, is_deleted=0).first()
    if not u:
        raise HTTPException(404, "User not found.")
    if u.role == "admin":
        raise HTTPException(400, "Admins already see every project.")

    wanted = {int(x) for x in project_ids.split(",") if x.strip().isdigit()}
    valid = {p.id for p in db.query(Project).filter(Project.is_active == 1).all()}
    wanted &= valid

    db.query(UserProject).filter(UserProject.user_id == user_id).delete()
    now = datetime.utcnow()
    for pid in sorted(wanted):
        db.add(UserProject(user_id=user_id, project_id=pid,
                           assigned_at=now, assigned_by=admin.username))

    # If they were last in a project they can no longer reach, forget it so
    # their next login resolves to one they can.
    if wanted and u.last_project_id not in wanted:
        u.last_project_id = None

    log_activity(
        db, user=admin, action="projects_assigned", target_type="user",
        target_id=user_id,
        details={"worker": u.username, "project_ids": sorted(wanted)},
    )
    db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)


@router.post("/users/create")
def create_user(
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form("worker"),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    username = username.strip()
    if not re.match(r"^[A-Za-z0-9_.-]{2,64}$", username):
        raise HTTPException(400, "Username must be 2–64 chars, alnum / _ . -")
    if len(password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters.")
    if role not in ("admin", "worker"):
        raise HTTPException(400, "Role must be admin or worker.")
    if db.query(User).filter_by(username=username).first():
        raise HTTPException(409, "Username already exists.")
    u = User(username=username, password_hash=hash_password(password), role=role)
    db.add(u)
    db.flush()
    log_activity(db, user=admin, action="user_created", target_type="user", target_id=u.id,
                 details={"username": username, "role": role})
    db.commit()
    return RedirectResponse("/admin/users", status_code=302)


@router.post("/users/{user_id}/toggle")
def toggle_user(user_id: int, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    u = db.query(User).filter_by(id=user_id).first()
    if not u:
        raise HTTPException(404, "User not found")
    if u.id == admin.id:
        raise HTTPException(400, "Cannot disable your own account.")
    u.is_active = 0 if u.is_active else 1
    log_activity(db, user=admin, action="user_toggled", target_type="user", target_id=u.id,
                 details={"is_active": bool(u.is_active)})
    db.commit()
    return RedirectResponse("/admin/users", status_code=302)


@router.post("/users/{user_id}/reset_password")
def reset_password(
    user_id: int,
    new_password: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if len(new_password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters.")
    u = db.query(User).filter_by(id=user_id).first()
    if not u:
        raise HTTPException(404, "User not found")
    u.password_hash = hash_password(new_password)
    log_activity(db, user=admin, action="password_reset", target_type="user", target_id=u.id)
    db.commit()
    return RedirectResponse("/admin/users", status_code=302)


@router.post("/users/{user_id}/delete")
def delete_user(
    user_id: int,
    confirm_username: str = Form(...),
    admin_password:   str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Soft-delete a user. Required confirmations: caller types the target
    user's username AND re-enters their own admin password.

    What this does:
      - Sets is_deleted=1, deleted_at=now on the User row. The user can no
        longer log in (auth filters is_deleted=0).
      - Reattributes the user's saved_posters.username to "[deleted: X]"
        so the admin gallery still shows them but no future query treats
        the original name as live.
      - Releases any titles the user had claimed (status → pending).
      - Keeps chat history and payment history intact (they reference the
        user's id, which still exists).

    Guards:
      - Cannot delete yourself.
      - Cannot delete the only remaining admin.
    """
    from ..auth import verify_password
    if not verify_password(admin_password, admin.password_hash):
        raise HTTPException(400, "Admin password is incorrect.")

    target = db.query(User).filter_by(id=user_id).first()
    if not target:
        raise HTTPException(404, "User not found.")
    if target.is_deleted:
        raise HTTPException(400, "User is already deleted.")
    if target.id == admin.id:
        raise HTTPException(400, "You can't delete your own account.")
    if confirm_username.strip() != target.username:
        raise HTTPException(
            400,
            f"Username confirmation didn't match. Expected: {target.username!r}",
        )
    if target.role == "admin":
        # Don't allow deleting the only remaining (active+non-deleted) admin.
        remaining_admins = (
            db.query(User)
              .filter(User.role == "admin",
                      User.is_active == 1,
                      User.is_deleted == 0,
                      User.id != target.id)
              .count()
        )
        if remaining_admins == 0:
            raise HTTPException(400, "Cannot delete the only remaining admin.")

    placeholder = f"[deleted: {target.username}]"
    now = datetime.utcnow()

    # Soft-delete the user row.
    target.is_deleted  = 1
    target.deleted_at  = now
    target.is_active   = 0
    target.locked_master_id = None

    # Reattribute saved-poster username for visual clarity. The user_id FK
    # remains correct so existing queries still join.
    db.query(SavedPoster).filter(SavedPoster.user_id == target.id) \
      .update({SavedPoster.username: placeholder}, synchronize_session=False)

    # Release any titles claimed by this user — pending for someone else.
    claims = (
        db.query(MasterTitle)
          .filter(MasterTitle.claimed_by_id == target.id,
                  MasterTitle.status == "in_progress")
          .all()
    )
    released = 0
    for t in claims:
        t.claimed_by_id   = None
        t.claimed_by_name = None
        t.claimed_at      = None
        t.status          = "pending"
        t.updated_at      = now
        released += 1

    log_activity(
        db, user=admin, action="user_deleted", target_type="user", target_id=target.id,
        details={"username": target.username, "released_claims": released,
                 "role": target.role},
    )
    db.commit()
    return JSONResponse({"ok": True, "released_claims": released})


@router.post("/users/{user_id}/release_queue")
def release_user_queue(
    user_id: int,
    keep_started: int = Form(1),  # default: only release un-started
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin override: release a user's claimed-but-unworked rows back to the pool.
    keep_started=1 (default) protects rows where work has begun (started_at set).
    keep_started=0 releases EVERYTHING back to pending, including started rows
    (their saved posters remain on disk; the title returns to the pool for someone else).
    """
    u = db.query(User).filter_by(id=user_id).first()
    if not u:
        raise HTTPException(404, "User not found")
    q = db.query(MasterTitle).filter(
        MasterTitle.claimed_by_id == user_id,
        MasterTitle.status == "in_progress",
    )
    if keep_started:
        q = q.filter(MasterTitle.started_at.is_(None))
    rows = q.all()
    now = datetime.utcnow()
    released_ids = []
    for r in rows:
        r.claimed_by_id = None
        r.claimed_by_name = None
        r.claimed_at = None
        r.status = "pending"
        r.updated_at = now
        released_ids.append(r.id)
    if u.locked_master_id in released_ids:
        u.locked_master_id = None
    log_activity(
        db, user=admin, action="released", target_type="user", target_id=user_id,
        details={"count": len(released_ids), "keep_started": bool(keep_started)},
    )
    db.commit()
    return JSONResponse({"ok": True, "released": len(released_ids)})


# ── Master sheet ─────────────────────────────────────────────────────────────

@router.get("/master", response_class=HTMLResponse)
def master_page(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(MASTER_PAGE_SIZE, ge=10, le=500),
    q: str = Query(""),
    status: str = Query(""),
    content_type: str = Query(""),
    needs_revision: int = Query(0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Initial render of the master page; the JS app pages further via /admin/api/master."""
    proj = current_project(request, admin, db)
    return templates.TemplateResponse(
        request,
        "admin_master.html",
        {"user": admin, "admin": admin,
            "page": page, "page_size": page_size, "q": q, "status": status,
            "content_type": content_type, "needs_revision": needs_revision,
            # A one-column artist sheet has no year and no movie/tv split, so
            # those column headers and the TYPE filter are dead controls.
            # Declared by the project rather than branched on its slug.
            "has_year": bool(proj.has_year),
            "has_content_type": bool(proj.has_content_type),
            "item_noun": proj.item_noun,
            "item_nouns": proj.item_noun_plural,
            "active_tab": "master",
        },
    )


def _master_query(db: Session, q: str, status: str, content_type: str, needs_revision: int,
                  project: Optional[Project] = None):
    """
    The single funnel for both the Title List page and its JSON API.

    `project` is not optional in practice — every caller passes the active
    project. It defaults to None only so the signature stays honest about
    what an unscoped query would mean (all projects), which is what the
    diagnostic tooling wants and what a project page must never get.
    """
    query = scope_titles(db.query(MasterTitle), project)
    if status in ("pending", "in_progress", "complete", "complete_pending", "skipped"):
        query = query.filter(MasterTitle.status == status)
    if content_type:
        query = query.filter(MasterTitle.content_type == content_type)
    if q.strip():
        like = f"%{q.strip()}%"
        query = query.filter(MasterTitle.title.ilike(like))
    if needs_revision:
        query = query.filter(MasterTitle.needs_revision == 1)
    return query


@router.get("/api/master")
def api_master(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(MASTER_PAGE_SIZE, ge=10, le=500),
    q: str = Query(""),
    status: str = Query(""),
    content_type: str = Query(""),
    needs_revision: int = Query(0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    query = _master_query(db, q, status, content_type, needs_revision,
                          project=current_project(request, admin, db))
    total = query.count()
    rows = (
        query
        .order_by(MasterTitle.external_id.asc().nullslast(), MasterTitle.id.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return JSONResponse({
        "page": page, "page_size": page_size, "total": total,
        "pages": (total + page_size - 1) // page_size,
        "items": [
            {
                "id": r.id, "external_id": r.external_id,
                "title": r.title, "year": r.year, "content_type": r.content_type,
                "votes": r.votes, "rating": r.rating,
                "status": r.status, "needs_revision": bool(r.needs_revision),
                "claimed_by": r.claimed_by_name,
                "skip_reason": r.skip_reason or "",
                "started": t_started_str(r),
                "description": (r.description or "")[:200],
            }
            for r in rows
        ],
    })


def t_started_str(r: MasterTitle) -> str:
    return r.original_save_date.isoformat() if r.original_save_date else ""


@router.post("/master/{title_id}/status")
def master_set_status(
    title_id: int,
    status: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin override: set a master's status directly.

    When admin sets status to 'complete', this also:
      - Resolves any open / awaiting-approval revisions on this master's
        posters (so the FLAG pill goes away).
      - Preserves the existing claim attribution (claimed_by stays as the
        last worker who actually did the work — admin doesn't take credit).
      - Clears any pending admin_note (the issue is being closed).
      - Clears the claiming user's locked_master_id if this was their lock.

    Going to 'pending' clears the claim. Other transitions leave the claim
    intact.
    """
    if status not in ("pending", "in_progress", "complete", "skipped"):
        raise HTTPException(400, "Bad status.")
    t = db.query(MasterTitle).filter_by(id=title_id).first()
    if not t:
        raise HTTPException(404, "Title not found.")
    old = t.status
    t.status = status
    resolved_revs = 0

    if status == "complete":
        # Resolve all active revisions on posters of this master, INCLUDING
        # revisions on soft-deleted posters (pending deletions count). The
        # old deleted_at filter was a defensive guard from when delete
        # auto-resolved; with the round-11 rework, delete leaves revisions
        # in awaiting_approval state.
        active = (
            db.query(Revision)
              .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
              .filter(
                  SavedPoster.master_title_id == t.id,
                  Revision.status.in_(("open", "awaiting_approval")),
              )
              .all()
        )
        for r in active:
            r.status = "resolved"
            r.resolved_by = admin.username
            r.resolved_at = datetime.utcnow()
            r.admin_verdict = "force-completed by admin"
            resolved_revs += 1
        t.needs_revision = 0
        t.completed_at = datetime.utcnow()
        t.admin_note = None
        # Free the claiming user's lock if it pointed here.
        if t.claimed_by_id:
            u = db.query(User).filter_by(id=t.claimed_by_id).first()
            if u and u.locked_master_id == t.id:
                u.locked_master_id = None

    elif status == "pending":
        t.claimed_by_id = None
        t.claimed_by_name = None
        t.claimed_at = None

    log_activity(
        db, user=admin, action="status_changed", target_type="master_title", target_id=t.id,
        details={"from": old, "to": status, "by_admin": True,
                 "resolved_revs": resolved_revs},
    )
    db.commit()
    return JSONResponse({"ok": True, "resolved_revs": resolved_revs})


@router.post("/master/bulk_status")
def master_bulk_status(
    request: Request,
    ids: str = Form(...),
    status: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if status not in ("pending", "complete", "skipped"):
        raise HTTPException(400, "Bad status for bulk operation.")
    try:
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    except ValueError:
        raise HTTPException(400, "Bad ids.")
    if not id_list:
        raise HTTPException(400, "No ids supplied.")

    # Scoped as well as filtered by id — the id list comes from the browser
    # and a bulk action must not reach outside the project it was fired from.
    rows = scope_titles(
        db.query(MasterTitle).filter(MasterTitle.id.in_(id_list)),
        current_project(request, admin, db),
    ).all()
    now = datetime.utcnow()
    total_resolved_revs = 0
    for r in rows:
        r.status = status
        r.updated_at = now
        if status == "pending":
            r.claimed_by_id = None
            r.claimed_by_name = None
            r.claimed_at = None
        elif status == "complete":
            r.completed_at = now
            # Resolve active flags so FLAG pill clears (same as single-row path).
            # We INCLUDE revisions on soft-deleted posters now — the round-11
            # rework changed delete-on-flagged from auto-resolve to admin-
            # pending, so deleted posters can carry awaiting_approval revisions.
            active = (
                db.query(Revision)
                  .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
                  .filter(
                      SavedPoster.master_title_id == r.id,
                      Revision.status.in_(("open", "awaiting_approval")),
                  )
                  .all()
            )
            for rev in active:
                rev.status = "resolved"
                rev.resolved_by = admin.username
                rev.resolved_at = now
                rev.admin_verdict = "force-completed by admin (bulk)"
                total_resolved_revs += 1
            r.needs_revision = 0
            r.admin_note = None
            if r.claimed_by_id:
                u = db.query(User).filter_by(id=r.claimed_by_id).first()
                if u and u.locked_master_id == r.id:
                    u.locked_master_id = None
    log_activity(
        db, user=admin, action="bulk_status", target_type="bulk", details={
            "count": len(rows), "ids": [r.id for r in rows], "status": status,
            "resolved_revs": total_resolved_revs,
        },
    )
    db.commit()
    return JSONResponse({"ok": True, "updated": len(rows),
                         "resolved_revs": total_resolved_revs})


@router.post("/master/clear")
def master_clear(request: Request, admin: User = Depends(require_admin),
                 db: Session = Depends(get_db)):
    """
    Wipe THIS PROJECT's title list.

    Scoped, not global: the button lives on a project page and clearing one
    niche must not take the others with it. Saved posters keep their data —
    their master_title_id dangles, but SavedPoster carries the immutable
    folder path so the files are still locatable.
    """
    proj = current_project(request, admin, db)
    q = scope_titles(db.query(MasterTitle), proj)
    n = q.count()
    scope_titles(db.query(MasterTitle), proj).delete(synchronize_session=False)
    log_activity(db, user=admin, action="bulk_status", target_type="bulk",
                 details={"deleted_master_rows": n, "project": proj.slug})
    db.commit()
    return RedirectResponse("/admin/master", status_code=302)


# ── Background master import ────────────────────────────────────────────────

IMPORT_LOCK = threading.Lock()


def _import_worker(job_id: int, raw_bytes: bytes, file_ext: str, replace: bool,
                   started_by: str, project_id: int):
    """
    Runs in a background thread. Owns its own DB session.

    `project_id` is which project's title list this file becomes. Two things
    depend on it and both are destructive to get wrong:

      * Imported rows are stamped with it. Rows used to be created with
        project_id NULL, which the whole app reads as "the default project" —
        so a celebrity sheet imported without this would silently join the
        movie queue.
      * REPLACE deletes only that project's rows. It used to delete the entire
        master_titles table. Importing a 2,000-row celebrity sheet with
        Replace ticked would have destroyed all 101,605 movie titles.
    """
    db = SessionLocal()
    try:
        job = db.query(ImportJob).filter_by(id=job_id).first()
        if not job:
            return
        job.state = "running"
        db.commit()

        # Parse rows
        rows = []
        if file_ext in (".xlsx", ".xlsm"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
            except Exception as e:
                job.state = "error"
                job.error = f"Excel read failed: {e}"
                job.finished_at = datetime.utcnow()
                db.commit()
                return
        else:
            text = raw_bytes.decode("utf-8-sig", errors="replace")
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            for r in reader:
                rows.append({(k or "").strip().lower(): v for k, v in r.items()})

        job.total_rows = len(rows)
        db.commit()

        if replace:
            # Scoped delete — see the docstring. NULL is folded in only when
            # importing into the default project, since that's what NULL means.
            from ..pipeline import DEFAULT_PROJECT_SLUG
            from ..models import Project as _Project
            default_row = db.query(_Project.id).filter_by(slug=DEFAULT_PROJECT_SLUG).first()
            is_default = bool(default_row) and default_row[0] == project_id
            q = db.query(MasterTitle).filter(MasterTitle.project_id == project_id)
            if is_default:
                q = db.query(MasterTitle).filter(
                    or_(MasterTitle.project_id == project_id,
                        MasterTitle.project_id.is_(None))
                )
            q.delete(synchronize_session=False)
            db.commit()

        # ── Which column holds the title? ───────────────────────────────
        # Sheets differ per niche and always will: one may have a `title`
        # column, another exactly one column called something else. Rather
        # than hardcode a list that needs editing for every future niche,
        # fall back to "the only column there is" when `title` is absent —
        # a single-column sheet is unambiguous by definition — and only then
        # to a list of likely names.
        #
        # The candidate list is a convenience, not the mechanism. Anything
        # not on it still imports fine as a single-column sheet, or by
        # naming the column `title`.
        title_key = "title"
        if rows and "title" not in rows[0]:
            headers = [k for k in rows[0].keys() if k]
            if len(headers) == 1:
                title_key = headers[0]
            else:
                for candidate in ("location", "place", "destination",
                                  "name", "subject"):
                    if candidate in rows[0]:
                        title_key = candidate
                        break

        # Insert in batches for speed
        batch = []
        BATCH_SIZE = 1000
        row_number = 0
        for r in rows:
            title = (str(r.get(title_key) or "")).strip()
            if not title:
                continue
            row_number += 1
            # external_id: try the literal "0" column (which is named "0" in our CSV header), then "num"
            ext = r.get("0") or r.get("num") or r.get("external_id")
            try:
                ext_id = int(str(ext).strip()) if ext not in (None, "") else None
            except (ValueError, TypeError):
                ext_id = None
            if ext_id is None:
                # No id column — number by position, 1-based, skipping blanks.
                # external_id is the universal join key: it prefixes the title
                # folder on disk and matches processed files back to posters,
                # so every row needs one and it must be stable for the life of
                # the sheet. Re-importing the SAME file reproduces the same
                # numbers; importing a re-ordered file would not, which is why
                # Replace exists rather than merging.
                ext_id = row_number
            year_raw = r.get("releaseyear") or r.get("release_year") or r.get("year") or ""
            year_str = str(year_raw).strip() if year_raw not in (None, "") else "N/A"
            m = re.search(r"\d{4}", year_str)
            year_str = m.group() if m else "N/A"
            try:
                votes = int(str(r.get("votes")).strip()) if r.get("votes") not in (None, "") else None
            except (ValueError, TypeError):
                votes = None
            try:
                rating = float(str(r.get("rating")).strip()) if r.get("rating") not in (None, "") else None
            except (ValueError, TypeError):
                rating = None
            content_type_raw = r.get("contenttype") or r.get("content_type") or None
            content_type = str(content_type_raw).strip() if content_type_raw else None
            description = r.get("description")
            description = str(description).strip() if description not in (None, "") else None

            mt = MasterTitle(
                external_id=ext_id, title=title, year=year_str,
                content_type=content_type, votes=votes, rating=rating,
                description=description, status="pending",
                project_id=project_id,
            )
            batch.append(mt)
            if len(batch) >= BATCH_SIZE:
                db.bulk_save_objects(batch)
                job.done_rows += len(batch)
                db.commit()
                batch = []

        if batch:
            db.bulk_save_objects(batch)
            job.done_rows += len(batch)
            db.commit()

        job.state = "done"
        job.finished_at = datetime.utcnow()
        db.commit()

        # Audit at the end (reuse the existing session from this worker thread)
        log_activity(
            db, user=None, action="imported", target_type="import_job", target_id=job.id,
            details={"by": started_by, "rows": job.done_rows,
                     "replaced": bool(replace), "project_id": project_id},
            commit=True,
        )
    except Exception as e:
        try:
            job = db.query(ImportJob).filter_by(id=job_id).first()
            if job:
                job.state = "error"
                job.error = str(e)
                job.finished_at = datetime.utcnow()
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


@router.post("/master/upload")
async def master_upload(
    request: Request,
    file: UploadFile = File(...),
    replace: int = Form(0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Empty file.")
    name = (file.filename or "").lower()
    ext = ".xlsx" if name.endswith(".xlsx") or name.endswith(".xlsm") else ".csv"

    job = ImportJob(
        started_by=admin.username, state="pending",
        total_rows=0, done_rows=0, replaced=int(bool(replace)),
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    # Resolved here, on the request thread, and passed in — the background
    # thread has no request and therefore no way to know which project the
    # admin was standing in.
    project = current_project(request, admin, db)

    threading.Thread(
        target=_import_worker,
        args=(job.id, raw, ext, bool(replace), admin.username, project.id),
        daemon=True,
    ).start()

    return JSONResponse({"ok": True, "job_id": job.id})


@router.get("/import/{job_id}")
def import_status(job_id: int, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    job = db.query(ImportJob).filter_by(id=job_id).first()
    if not job:
        raise HTTPException(404, "Unknown job.")
    return JSONResponse({
        "id": job.id, "state": job.state, "total": job.total_rows,
        "done": job.done_rows, "error": job.error or "",
        "replaced": bool(job.replaced),
        "started_by": job.started_by,
    })


# ── Image browser ────────────────────────────────────────────────────────────

@router.get("/browse", response_class=HTMLResponse)
def browse_page(
    request: Request,
    worker: Optional[str] = None,
    date: Optional[str] = None,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    # Resolved once and reused: the folder segment narrows the filesystem
    # scan to this project's tree, and the slug scopes the sticky default.
    proj = current_project(request, admin, db)
    from ..workspace_migration import project_folder_for
    proj_folder = project_folder_for(proj)

    workers = list_users_with_workspaces(proj_folder)
    if not workers:
        users = db.query(User).filter_by(role="worker").order_by(User.username.asc()).all()
        workers = [u.username for u in users]

    # Sticky default: if worker param given, save it as default for this admin.
    # If no param, read the saved default. Fallback to alphabetical first.
    # Scoped per project: an admin reviewing movies and celebrities has a
    # different "usual worker" in each, and a shared key would make switching
    # project silently change who you're looking at.
    settings_key = f"admin.{admin.id}.{proj.slug}.browse_default_worker"
    if worker:
        # Save this selection as the new default.
        setting = db.query(AppSetting).filter_by(key=settings_key).first()
        if setting:
            setting.value = worker
        else:
            db.add(AppSetting(key=settings_key, value=worker))
        db.commit()
        selected_worker = worker
    else:
        setting = db.query(AppSetting).filter_by(key=settings_key).first()
        saved = setting.value if setting else None
        # Only use the saved default if that worker still has a workspace.
        if saved and saved in workers:
            selected_worker = saved
        else:
            selected_worker = workers[0] if workers else ""

    dates = list_date_folders(selected_worker, proj_folder) if selected_worker else []
    today_iso = local_today().isoformat()
    if today_iso not in dates and selected_worker:
        dates = [today_iso] + dates
    selected_date = date or (today_iso if today_iso in dates else (dates[0] if dates else today_iso))

    return templates.TemplateResponse(
        request,
        "admin_image_browser.html",
        {"user": admin, "admin": admin, "workers": workers, "dates": dates,
            "selected_worker": selected_worker, "selected_date": selected_date,
            "active_tab": "browse",
        },
    )


@router.get("/api/browse")
def api_browse(
    request: Request,
    worker: str,
    date: str,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Group all live SavedPosters for a (user, original_save_date) pair by master title.
    Each title returns its full poster set, with revision info per poster.
    """
    try:
        d = date_type.fromisoformat(date)
    except ValueError:
        raise HTTPException(400, "Bad date.")

    rows = (
        db.query(SavedPoster, MasterTitle, Revision)
          .join(MasterTitle, SavedPoster.master_title_id == MasterTitle.id)
          .outerjoin(
              Revision,
              and_(
                  Revision.saved_poster_id == SavedPoster.id,
                  Revision.status.in_(("open", "awaiting_approval")),
              ),
          )
          .filter(
              SavedPoster.username == worker,
              SavedPoster.original_save_date == d,
              SavedPoster.deleted_at.is_(None),
          )
    )
    # Scoped through the joined MasterTitle rather than the poster: posters
    # carry no project of their own, and deliberately so — a title's project
    # is the one fact, and duplicating it onto every poster would be a second
    # copy to keep in sync through renames, replacements and reassignments.
    rows = (
        scope_titles(rows, current_project(request, admin, db))
          .order_by(SavedPoster.title_folder_path.asc(), SavedPoster.filename.asc())
          .all()
    )

    # Group by master title (preserve folder-name ordering)
    from collections import OrderedDict
    titles = OrderedDict()
    for sp, mt, rev in rows:
        key = mt.id if mt else None
        if key not in titles:
            titles[key] = {
                "master_id": key,
                "title": mt.title if mt else "(unknown)",
                "year":  mt.year  if mt else "",
                "title_folder": sp.title_folder_path,
                "needs_revision": bool(mt.needs_revision) if mt else False,
                "posters": [],
            }
        titles[key]["posters"].append({
            "poster_id": sp.id,
            "filename": sp.filename,
            "title_folder": sp.title_folder_path,
            "size": sp.file_size,
            "low_quality_url": bool(sp.low_quality_url),
            "image_width":  sp.image_width,
            "image_height": sp.image_height,
            "added_by":     sp.added_by or None,
            "flagged": rev is not None,
            "revision_id": rev.id if rev else None,
            "revision_status": rev.status if rev else None,
            "revision_type":   rev.revision_type if rev else None,
            "comment": rev.comment if rev else "",
            "worker_note": rev.worker_note if rev else "",
        })

    from ..pipeline import get_setting
    try:
        min_width = int(get_setting(db, "review_min_width_px",
                                    project=current_project(request, admin, db)) or 0)
    except Exception:
        min_width = 800

    return JSONResponse({
        "worker": worker, "date": date,
        "min_width": min_width,
        "title_count": len(titles),
        "poster_count": sum(len(t["posters"]) for t in titles.values()),
        "titles": list(titles.values()),
    })


@router.get("/api/poster/{poster_id}/timeline")
def poster_timeline(
    poster_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Everything that ever happened to one poster, oldest first.

    ════════════════════════════════════════════════════════════════════════
    WHY THIS EXISTS
    ════════════════════════════════════════════════════════════════════════
    The state of a poster is spread across six tables — the poster row, the
    activity log, revisions, processed_images, upload_tracking and the title
    it belongs to. Answering "why is this image not on FineArtAmerica?"
    currently means opening a SQLite shell and joining them by hand.

    That question gets asked constantly and will get asked more as niches and
    marketplaces multiply, so it deserves an answer in the UI. This assembles
    one ordered list from all six, in the operator's own vocabulary.

    Read-only, by design: this is a place to understand what happened, never
    to change it. Anything that mutates state belongs on the page that owns
    that state, where the confirmation and the audit entry already exist.
    """
    sp = db.query(SavedPoster).filter_by(id=poster_id).first()
    if not sp:
        raise HTTPException(404, "Poster not found.")
    title = db.query(MasterTitle).filter_by(id=sp.master_title_id).first()

    events: list[dict] = []

    def add(when, kind: str, text: str, detail: str = "", actor: str = ""):
        if when is None:
            return
        events.append({
            "at": when.isoformat() if hasattr(when, "isoformat") else str(when),
            "kind": kind, "text": text, "detail": detail, "actor": actor,
        })

    # ── Sourcing ────────────────────────────────────────────────────────────
    add(sp.created_at, "saved",
        "Downloaded by " + (sp.added_by or sp.username),
        sp.source_url or "", sp.added_by or sp.username)
    if sp.low_quality_url:
        add(sp.created_at, "warn", "Saved past the low-quality warning",
            f"{sp.image_width or '?'}×{sp.image_height or '?'}", sp.username)

    # ── Admin/worker actions from the audit log ─────────────────────────────
    for row in (
        db.query(ActivityLog)
          .filter(ActivityLog.target_type == "saved_poster",
                  ActivityLog.target_id == poster_id)
          .order_by(ActivityLog.created_at.asc())
          .all()
    ):
        add(row.created_at, row.action, row.action.replace("_", " ").title(),
            row.details or "", row.username or "system")

    # ── Change requests ─────────────────────────────────────────────────────
    for rev in (
        db.query(Revision)
          .filter(Revision.saved_poster_id == poster_id)
          .order_by(Revision.created_at.asc())
          .all()
    ):
        add(rev.created_at, "flagged", "Changes requested", rev.comment or "")
        add(rev.submitted_at, "resubmitted", "Worker submitted a fix", rev.worker_note or "")
        add(rev.resolved_at, "resolved", f"Change request {rev.status}",
            rev.admin_verdict or "")

    # ── Pipeline: greenlight, Photoshop, upload ─────────────────────────────
    if title is not None and title.greenlit_at:
        src = title.greenlit_source or "unknown"
        add(title.greenlit_at, "greenlit",
            "Released into the pipeline",
            f"source: {src}", title.greenlit_by or "")

    for pi in (
        db.query(ProcessedImage)
          .filter(ProcessedImage.saved_poster_id == poster_id)
          .order_by(ProcessedImage.created_at.asc())
          .all()
    ):
        add(pi.created_at, "processed",
            "Processed" + ("" if pi.is_current else " (superseded)"),
            f"{pi.storage_path} · {round((pi.duration_ms or 0) / 1000)}s"
            + (f" · script {pi.script_version}" if pi.script_version else ""),
            pi.processed_by or "")

    for ut, acct in (
        db.query(UploadTracking, UploadAccount)
          .outerjoin(UploadAccount, UploadTracking.account_id == UploadAccount.id)
          .filter(UploadTracking.saved_poster_id == poster_id)
          .all()
    ):
        who = acct.name if acct else f"account {ut.account_id}"
        add(ut.created_at, "upload_queued", f"Queued for {who}", ut.target_site or "")
        add(ut.uploaded_at, "uploaded", f"Live on {who}", ut.remote_title or "")
        add(ut.removed_at, "removed", f"Removed from {who}", ut.removed_reason or "")
        if ut.status == "failed" and ut.last_error:
            # UploadTracking keeps no per-attempt timestamp, so the failure is
            # pinned to the claim that produced it. Close enough to order the
            # timeline correctly, and honest about what's actually recorded.
            add(ut.claimed_at or ut.created_at, "failed",
                f"Upload failed on {who} (attempt {ut.attempts})",
                ut.last_error[:400])

    if sp.process_error:
        add(sp.claimed_at or sp.created_at, "failed", "Processing failed",
            sp.process_error[:400], sp.claimed_by or "")

    add(sp.deleted_at, "deleted", "Deleted", sp.delete_note or "")

    events.sort(key=lambda e: e["at"])

    return JSONResponse({
        "poster": {
            "id": sp.id,
            "filename": sp.filename,
            "folder": sp.title_folder_path,
            "worker": sp.username,
            "saved_on": sp.original_save_date.isoformat() if sp.original_save_date else "",
            "size": sp.file_size,
            "dimensions": (f"{sp.image_width}×{sp.image_height}"
                           if sp.image_width and sp.image_height else ""),
            "pipeline_status": sp.pipeline_status or "not greenlit",
            "source_url": sp.source_url,
            "deleted": sp.deleted_at is not None,
        },
        "title": {
            "id": title.id if title else None,
            "name": title.title if title else "(unknown)",
            "year": title.year if title else "",
            "external_id": title.external_id if title else None,
        },
        "events": events,
    })


@router.get("/file/{poster_id}")
def serve_poster_file(
    poster_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    sp = db.query(SavedPoster).filter_by(id=poster_id).first()
    if not sp or sp.deleted_at is not None:
        raise HTTPException(404, "File not found.")
    p = saved_poster_path(sp)
    if not p.is_file():
        raise HTTPException(404, "File missing on disk.")
    return FileResponse(p)


# ── Admin poster management (browse page delete/upload) ──────────────────────

@router.post("/poster/{poster_id}/delete")
def admin_delete_poster(
    poster_id: int,
    note: str = Form(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin hard-deletes a poster. Does NOT count against the worker.
    Used for cleanup — e.g. removing a poster that causes inconsistency,
    or cleaning up test data. Resolves any active revisions on it.
    """
    sp = db.query(SavedPoster).filter_by(id=poster_id).first()
    if not sp:
        raise HTTPException(404, "Poster not found.")
    # Remove file from disk
    fs_path = saved_poster_path(sp)
    fs_path.unlink(missing_ok=True)
    # Resolve any active revisions
    revs = (
        db.query(Revision)
          .filter(Revision.saved_poster_id == sp.id,
                  Revision.status.in_(("open", "awaiting_approval")))
          .all()
    )
    for r in revs:
        r.status = "resolved"
        r.resolved_by = admin.username
        r.resolved_at = datetime.utcnow()
        r.admin_verdict = "admin-deleted" + (f": {note.strip()}" if note.strip() else "")
    # Soft-delete the poster
    sp.deleted_at = datetime.utcnow()
    sp.delete_note = f"[admin] {note.strip()}" if note.strip() else "[admin] deleted by admin"
    # Recompute title state
    mt = db.query(MasterTitle).filter_by(id=sp.master_title_id).first()
    if mt:
        remaining = (
            db.query(SavedPoster)
              .filter_by(master_title_id=mt.id)
              .filter(SavedPoster.deleted_at.is_(None))
              .count()
        )
        any_active = (
            db.query(Revision)
              .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
              .filter(SavedPoster.master_title_id == mt.id,
                      Revision.status.in_(("open", "awaiting_approval")))
              .count()
        )
        mt.needs_revision = 1 if any_active else 0
        if remaining == 0 and mt.status in ("in_progress", "complete_pending", "complete", "skipped"):
            mt.status = "pending"
            mt.needs_revision = 0
            mt.completed_at = None
            if mt.claimed_by_id:
                u = db.query(User).filter_by(id=mt.claimed_by_id).first()
                if u and u.locked_master_id == mt.id:
                    u.locked_master_id = None
                mt.claimed_by_id = None
            # Clean up empty folder
            title_dir = fs_path.parent
            if title_dir.is_dir() and not list(title_dir.iterdir()):
                import shutil
                shutil.rmtree(title_dir, ignore_errors=True)

    log_activity(
        db, user=admin, action="admin_deleted", target_type="saved_poster", target_id=sp.id,
        details={"filename": sp.filename, "master_id": sp.master_title_id,
                 "note": note.strip() or None},
    )
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/poster/add")
def admin_add_poster(
    master_id: int = Form(...),
    url: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin adds a poster to a title by URL. The poster is saved under the
    title's existing worker (claimed_by), or under 'admin' if unclaimed.
    Does NOT count toward the worker's save stats or payment — admin's
    responsibility to manage. Useful for fixing titles that need a poster
    the worker missed.
    """
    from .worker import (
        _validate_image_url, _download_to, _ensure_first_save_metadata,
        _is_low_quality_url,
    )
    from ..utils import title_folder_for, count_live_posters_for_master
    from ..parsing import filename_for
    from ..imghdr_lite import read_file_dimensions

    t = db.query(MasterTitle).filter_by(id=master_id).first()
    if not t:
        raise HTTPException(404, "Title not found.")

    from ..pipeline import resolve_project

    ok, reason = _validate_image_url(
        url.strip(), db, resolve_project(db, t.project_id))
    if not ok:
        raise HTTPException(400, reason)
    src_url = url.strip()

    # Determine the worker username for folder placement.
    worker_username = "admin"
    worker_id = admin.id
    if t.claimed_by_id:
        claimer = db.query(User).filter_by(id=t.claimed_by_id).first()
        if claimer:
            worker_username = claimer.username
            worker_id = claimer.id

    today = local_today()
    _ensure_first_save_metadata(t, today)
    db.flush()

    # Same project stamping as the worker save path — an admin-added image
    # must land in the same tree as the worker's own.
    from ..workspace_migration import project_folder_for
    project_folder = project_folder_for(ensure_default_project(db)
                                        if not t.project_id else
                                        db.query(Project).filter_by(id=t.project_id).first())

    folder = title_folder_for(worker_username, t.original_save_date,
                              t.title_folder_path, project_folder)

    base = count_live_posters_for_master(db, t.id)
    count = base + 1
    target_name = filename_for(t.title, count, src_url)
    target_path = folder / target_name
    while target_path.exists():
        count += 1
        target_name = filename_for(t.title, count, src_url)
        target_path = folder / target_name

    written = _download_to(src_url, target_path)
    dims = read_file_dimensions(target_path)
    img_w, img_h = (dims if dims else (None, None))

    sp = SavedPoster(
        master_title_id    = t.id,
        user_id            = worker_id,
        username           = worker_username,
        project_folder     = project_folder,
        original_save_date = t.original_save_date,
        title_folder_path  = t.title_folder_path,
        filename           = target_name,
        source_url         = src_url,
        file_size          = written,
        low_quality_url    = 0,
        image_width        = img_w,
        image_height       = img_h,
        added_by           = admin.username,
    )
    db.add(sp)
    db.flush()  # assign sp.id before logging

    # If title was pending/skipped, move to in_progress.
    if t.status in ("pending", "skipped"):
        t.status = "in_progress"
        t.skip_reason = None
        if not t.claimed_by_id:
            t.claimed_by_id = admin.id

    log_activity(
        db, user=admin, action="admin_added", target_type="saved_poster", target_id=sp.id,
        details={"filename": target_name, "master_id": t.id,
                 "url": src_url, "title": f"{t.title} ({t.year})"},
    )
    db.commit()
    return JSONResponse({"ok": True, "filename": target_name, "poster_id": sp.id})


# ── Revisions ────────────────────────────────────────────────────────────────

@router.post("/poster/{poster_id}/flag")
def flag_poster(
    poster_id: int,
    comment: str = Form(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    sp = db.query(SavedPoster).filter_by(id=poster_id).first()
    if not sp or sp.deleted_at is not None:
        raise HTTPException(404, "Poster not found.")
    if sp.added_by:
        raise HTTPException(400, "This poster was added by admin — it cannot be flagged for worker revision.")

    # If there's already an active revision (open OR awaiting_approval), update it
    # back to 'open' with the new comment instead of stacking another row.
    existing = (
        db.query(Revision)
          .filter(
              Revision.saved_poster_id == sp.id,
              Revision.status.in_(("open", "awaiting_approval")),
          )
          .first()
    )
    if existing:
        existing.comment       = comment.strip() or None
        existing.flagged_by    = admin.username
        existing.status        = "open"
        existing.submitted_at  = None
        existing.worker_note   = None
        existing.admin_verdict = None
        log_activity(db, user=admin, action="flagged", target_type="revision", target_id=existing.id,
                     details={"updated": True, "poster_id": sp.id})
        db.commit()
        return JSONResponse({"ok": True, "id": existing.id, "updated": True})

    mt = db.query(MasterTitle).filter_by(id=sp.master_title_id).first()
    if mt:
        mt.needs_revision = 1

    rev = Revision(
        saved_poster_id=sp.id,
        comment=comment.strip() or None,
        flagged_by=admin.username,
        status="open",
    )
    db.add(rev)
    db.flush()
    log_activity(db, user=admin, action="flagged", target_type="revision", target_id=rev.id,
                 details={"poster_id": sp.id, "filename": sp.filename})
    db.commit()
    return JSONResponse({"ok": True, "id": rev.id, "updated": False})


@router.post("/poster/{poster_id}/unflag")
def unflag_poster(
    poster_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Clear all flags from a single poster. Three cases to handle:

      1. Simple revision where saved_poster_id = sp.id  → delete it.
      2. Similar revision where saved_poster_id = sp.id (i.e. sp is the
         primary participant) → delete it. The whole comparison goes away
         since the primary is no longer flagged.
      3. Similar revision where sp.id is in `related_poster_ids` JSON but
         NOT the primary → remove sp from the list. If the list shrinks
         to <2 entries (a "similar pair" of one is meaningless), delete
         the whole revision.

    Without case 3, clicking CLEAR FLAG on a non-primary participant of a
    similar revision would leave the revision active, the master flagged,
    and the worker stuck (they could only clear by acting on the primary).
    """
    import json as _json
    sp = db.query(SavedPoster).filter_by(id=poster_id).first()
    if not sp:
        raise HTTPException(404, "Poster not found.")

    # Cases 1+2 — direct revisions on this poster.
    direct = (
        db.query(Revision)
          .filter(
              Revision.saved_poster_id == sp.id,
              Revision.status.in_(("open", "awaiting_approval")),
          )
          .all()
    )
    cleared = 0
    for r in direct:
        db.delete(r)
        cleared += 1

    # Case 3 — similar revisions where sp is in related_poster_ids but
    # not the primary. We have to scan all open similar revisions because
    # related_poster_ids is JSON-encoded text, not a relational column.
    sim_others = (
        db.query(Revision)
          .filter(
              Revision.status.in_(("open", "awaiting_approval")),
              Revision.revision_type == "similar",
              Revision.saved_poster_id != sp.id,  # primaries already handled above
          )
          .all()
    )
    for r in sim_others:
        try:
            related = _json.loads(r.related_poster_ids or "[]")
        except (TypeError, ValueError):
            related = []
        if sp.id not in related:
            continue
        new_related = [pid for pid in related if pid != sp.id]
        if len(new_related) < 2:
            # A similar revision needs ≥2 participants to make sense.
            db.delete(r)
        else:
            r.related_poster_ids = _json.dumps(new_related)
        cleared += 1

    # Recompute master.needs_revision: any active simple/similar revisions
    # tied to this master? Filter out soft-deleted posters from the JOIN
    # defensively (A3) — under normal flow their revisions would already
    # be resolved, but if anything ever wedged this prevents a stuck flag.
    mt = db.query(MasterTitle).filter_by(id=sp.master_title_id).first()
    if mt:
        any_active = (
            db.query(Revision)
              .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
              .filter(
                  SavedPoster.master_title_id == mt.id,
                  SavedPoster.deleted_at.is_(None),
                  Revision.status.in_(("open", "awaiting_approval")),
              )
              .count()
        )
        mt.needs_revision = 1 if any_active else 0

    log_activity(db, user=admin, action="unflagged", target_type="saved_poster", target_id=sp.id,
                 details={"cleared": cleared})
    db.commit()
    return JSONResponse({"ok": True, "cleared": cleared})


# ── Mark posters as too similar (special revision type) ─────────────────────

@router.post("/posters/mark_similar")
def mark_similar(
    poster_ids: str = Form(...),       # comma-separated list
    comment: str = Form(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin selects 2+ posters in the gallery and clicks "These are too similar".
    Creates ONE Revision of type='similar' linking all the selected posters.
    The worker sees them as a group and picks one to redo (REPLACE / DELETE).
    """
    import json as _json
    try:
        ids = [int(x) for x in poster_ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(400, "Bad poster_ids — must be a comma-separated list of integers.")
    ids = list(dict.fromkeys(ids))  # de-dup, preserve order
    if len(ids) < 2:
        raise HTTPException(400, "Pick at least two posters to mark as similar.")

    posters = (
        db.query(SavedPoster)
          .filter(SavedPoster.id.in_(ids), SavedPoster.deleted_at.is_(None))
          .all()
    )
    if len(posters) < 2:
        raise HTTPException(404, "Some selected posters no longer exist.")

    # Admin-added posters cannot be flagged for worker action.
    admin_added = [p for p in posters if p.added_by]
    if admin_added:
        names = ", ".join(p.filename for p in admin_added)
        raise HTTPException(400, f"Cannot include admin-added posters in similar marking: {names}")

    # All must belong to the same master title — comparing across titles is meaningless.
    master_ids = {p.master_title_id for p in posters}
    if len(master_ids) != 1:
        raise HTTPException(400, "All selected posters must belong to the same title.")
    # All must belong to the same user — the revision is for them to act on.
    user_ids = {p.user_id for p in posters}
    if len(user_ids) != 1:
        raise HTTPException(400, "All selected posters must belong to the same user.")

    # Use the first as the primary saved_poster_id; keep all in related_poster_ids
    # (including the primary, so the worker UI just iterates one list).
    primary = posters[0]
    rev = Revision(
        saved_poster_id    = primary.id,
        comment            = comment.strip() or "These two posters look too similar — pick one to replace.",
        flagged_by         = admin.username,
        status             = "open",
        revision_type      = "similar",
        related_poster_ids = _json.dumps([p.id for p in posters]),
    )
    db.add(rev)

    # Mark the master as needing revision so it surfaces on the worker side.
    mt = db.query(MasterTitle).filter_by(id=primary.master_title_id).first()
    if mt:
        mt.needs_revision = 1

    db.flush()
    log_activity(
        db, user=admin, action="flagged_similar", target_type="revision", target_id=rev.id,
        details={"poster_ids": ids, "master_id": primary.master_title_id},
    )
    db.commit()
    return JSONResponse({"ok": True, "revision_id": rev.id})


# ── Approval / rejection of awaiting_approval revisions ─────────────────────

@router.post("/revisions/{revision_id}/approve")
def approve_revision(
    revision_id: int,
    verdict: str = Form(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Admin confirms the worker's fix. Revision → resolved, flag clears from worker side."""
    rev = db.query(Revision).filter_by(id=revision_id).first()
    if not rev:
        raise HTTPException(404, "Revision not found.")
    if rev.status not in ("awaiting_approval", "open"):
        raise HTTPException(400, "Revision is not in a state that can be approved.")
    rev.status = "resolved"
    rev.resolved_by = admin.username
    rev.resolved_at = datetime.utcnow()
    rev.admin_verdict = verdict.strip() or None

    # Recompute needs_revision on the master
    sp = db.query(SavedPoster).filter_by(id=rev.saved_poster_id).first()
    if sp:
        mt = db.query(MasterTitle).filter_by(id=sp.master_title_id).first()
        if mt:
            any_active = (
                db.query(Revision)
                  .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
                  .filter(
                      SavedPoster.master_title_id == mt.id,
                      SavedPoster.deleted_at.is_(None),
                      Revision.status.in_(("open", "awaiting_approval")),
                      Revision.id != rev.id,
                  )
                  .count()
            )
            mt.needs_revision = 1 if any_active else 0

    log_activity(db, user=admin, action="approved", target_type="revision", target_id=rev.id,
                 details={"verdict": verdict.strip() or None})
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/revisions/{revision_id}/reject")
def reject_revision(
    revision_id: int,
    verdict: str = Form(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Admin says 'no, redo it'. Revision → open with new comment for worker."""
    rev = db.query(Revision).filter_by(id=revision_id).first()
    if not rev:
        raise HTTPException(404, "Revision not found.")
    if rev.status != "awaiting_approval":
        raise HTTPException(400, "Revision is not awaiting approval.")
    rev.status = "open"
    rev.admin_verdict = verdict.strip() or None
    # Keep submitted_at so the worker side can detect this is a re-opened-after-rejection.
    # The combination (status=open AND admin_verdict set) is the "REJECTED" indicator.
    # Append the verdict to the comment so the worker sees it as the new instruction
    if verdict.strip():
        prefix = (rev.comment or "").strip()
        rev.comment = (prefix + "\n\n[admin]: " + verdict.strip()) if prefix else ("[admin]: " + verdict.strip())
    log_activity(db, user=admin, action="rejected", target_type="revision", target_id=rev.id,
                 details={"verdict": verdict.strip() or None})
    db.commit()
    return JSONResponse({"ok": True})


# ── Approval / rejection of complete_pending titles ────────────────────────

@router.post("/title/{master_id}/approve_complete")
def approve_complete(
    master_id: int,
    verdict: str = Form(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin approves a complete_pending title. The title becomes 'complete'
    and every active revision on its posters (open or awaiting_approval)
    is resolved with the admin's verdict.

    This is the round-11 path when worker clicks DONE on a title that had
    flags or pending deletions. Replaces the round-9 force=1 bypass.
    """
    t = db.query(MasterTitle).filter_by(id=master_id).first()
    if not t:
        raise HTTPException(404, "Title not found.")
    if t.status != "complete_pending":
        raise HTTPException(400, f"Title is not pending — status is '{t.status}'.")

    # ── v15: Guard — cannot approve completion with zero live posters ──
    live_count = count_live_posters_for_master(db, t.id)
    if live_count == 0:
        raise HTTPException(
            400,
            "Cannot approve — this title has no live posters. "
            "Reject it back to the worker, or send it to Skipped."
        )

    now = datetime.utcnow()
    # Resolve ALL revisions on this title's posters (including deletes).
    revs = (
        db.query(Revision)
          .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
          .filter(
              SavedPoster.master_title_id == t.id,
              Revision.status.in_(("open", "awaiting_approval")),
          )
          .all()
    )
    resolved_ids = []
    suffix = (": " + verdict.strip()) if verdict.strip() else ""
    for r in revs:
        r.status = "resolved"
        r.resolved_by = admin.username
        r.resolved_at = now
        r.admin_verdict = "approved via title completion" + suffix
        resolved_ids.append(r.id)

    t.status = "complete"
    t.completed_at = now
    t.needs_revision = 0
    t.admin_note = None
    # Free the claiming user's lock if it pointed here (defensive — the
    # worker route already cleared it on submit, but a re-submission path
    # could leave one stale).
    if t.claimed_by_id:
        u = db.query(User).filter_by(id=t.claimed_by_id).first()
        if u and u.locked_master_id == t.id:
            u.locked_master_id = None

    log_activity(
        db, user=admin, action="approved_completion",
        target_type="master_title", target_id=t.id,
        details={"verdict": verdict.strip() or None,
                 "resolved_revisions": resolved_ids},
    )
    db.commit()
    return JSONResponse({"ok": True, "resolved": len(resolved_ids)})


@router.post("/title/{master_id}/reject_complete")
def reject_complete(
    master_id: int,
    verdict: str = Form(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin rejects a complete_pending title. The title goes back to
    'in_progress', all its awaiting_approval revisions revert to 'open'
    with the admin's verdict pinned (so worker sees what's wrong), and
    the admin's verdict is also pinned as an admin_note on the master
    title so the worker sees a top-level summary.

    A rejection requires a verdict — otherwise the worker has no idea
    what to do differently.
    """
    verdict = (verdict or "").strip()
    if not verdict:
        raise HTTPException(400, "A verdict / note is required when rejecting.")
    t = db.query(MasterTitle).filter_by(id=master_id).first()
    if not t:
        raise HTTPException(404, "Title not found.")
    if t.status != "complete_pending":
        raise HTTPException(400, f"Title is not pending — status is '{t.status}'.")

    now = datetime.utcnow()
    revs = (
        db.query(Revision)
          .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
          .filter(
              SavedPoster.master_title_id == t.id,
              Revision.status == "awaiting_approval",
          )
          .all()
    )
    reopened_ids = []
    for r in revs:
        r.status = "open"
        r.admin_verdict = verdict  # pin on the revision so worker sees per-flag context
        r.submitted_at = None
        # Keep worker_action set — it tells the worker what they did
        # previously (so the rejection card can say "your deletion was
        # rejected" or "your replacement was rejected").
        reopened_ids.append(r.id)

    t.status = "in_progress"
    t.completed_at = None
    t.admin_note = verdict
    t.needs_revision = 1 if reopened_ids else t.needs_revision
    # Re-lock the title to the worker so they can pick up where they were.
    if t.claimed_by_id:
        u = db.query(User).filter_by(id=t.claimed_by_id).first()
        if u and not u.locked_master_id:
            u.locked_master_id = t.id

    log_activity(
        db, user=admin, action="rejected_completion",
        target_type="master_title", target_id=t.id,
        details={"verdict": verdict, "reopened_revisions": reopened_ids},
    )
    db.commit()
    return JSONResponse({"ok": True, "reopened": len(reopened_ids)})


# ── Skip-revise: admin sends a skipped title back to the worker with a note ──

@router.post("/title/{master_id}/skip_revise")
def skip_revise(
    master_id: int,
    note: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin reviews a skipped title and tells the worker to take another look.
    Resets the title to in_progress (still claimed by original worker if they're
    still around; otherwise pending) and stores the admin's note for display.
    """
    t = db.query(MasterTitle).filter_by(id=master_id).first()
    if not t:
        raise HTTPException(404, "Title not found.")
    if t.status != "skipped":
        raise HTTPException(400, "Title is not skipped.")
    note = note.strip()
    if not note:
        raise HTTPException(400, "A note is required when sending a skipped title back.")

    # Preserve the skip_reason for the worker's reference; just add admin_note.
    t.admin_note = note
    if t.claimed_by_id is not None:
        # Original claimer is still around — bounce back to in_progress for them
        t.status = "in_progress"
    else:
        # Was unclaimed — push back to pending so anyone can pick it up
        t.status = "pending"
    log_activity(
        db, user=admin, action="skip_revised", target_type="master_title", target_id=t.id,
        details={"note": note, "previous_skip_reason": t.skip_reason or ""},
    )
    db.commit()
    return JSONResponse({"ok": True})


# `POST /title/{id}/clear_admin_note` was removed on 2026-08-24. It had no
# button, no form and no fetch anywhere on the site — it could only ever have
# been called by hand. An endpoint with zero callers is not a spare part; it
# is something a future session has to read and reason about before deciding
# it does nothing. The note is still cleared wherever it is actually edited.


@router.get("/revisions", response_class=HTMLResponse)
def revisions_page(
    request: Request,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin's central place to handle worker pushback.

    Round-11 layout (top → bottom):

      1. PENDING COMPLETIONS — titles in 'complete_pending'. Worker
         clicked DONE while there were changes; admin must approve the
         whole title at once. Each card shows the title's diff (all
         awaiting_approval revisions inline) so admin sees what changed.

      2. AWAITING YOUR APPROVAL — revisions in 'awaiting_approval' whose
         title is NOT in 'complete_pending' (deduped). These are
         standalone "worker fixed the flag but didn't try to complete
         the title yet" cases.

      3. OPEN — WAITING ON USER — revisions in 'open'. Admin flagged,
         worker hasn't acted. Unchanged from before.

      4. RECENT MISTAKE DELETIONS — auto-resolved deletions on
         non-flagged posters. After the round-11 rework, this section
         only catches "worker downloaded the wrong image and deleted it"
         cases. (Flagged-poster deletions now go through pending
         completions or awaiting approval instead.)

      5. RESOLVED — history.
    """
    # Every section on this page is scoped to the project the admin is inside.
    # Revisions hang off posters, and posters hang off titles, so the join to
    # MasterTitle is what carries the project — which is why each query below
    # joins it even where the section doesn't display title fields.
    proj = current_project(request, admin, db)

    # ── Pending completions: titles in complete_pending ─────────────────────
    pending_titles = (
        scope_titles(db.query(MasterTitle), proj)
          .filter(MasterTitle.status == "complete_pending")
          .order_by(MasterTitle.updated_at.desc().nullslast())
          .all()
    )
    pending_complete_blocks = []
    pending_title_ids = set()
    for t in pending_titles:
        # All active revisions on this title's posters (open or awaiting).
        revs_for_title = (
            db.query(Revision, SavedPoster)
              .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
              .filter(
                  SavedPoster.master_title_id == t.id,
                  Revision.status.in_(("open", "awaiting_approval")),
              )
              .order_by(Revision.created_at.asc())
              .all()
        )
        # Activity log entries that show what the worker did since the title
        # was last in 'pending' or 'in_progress' — additions, deletions,
        # replacements. Used by admin to scan a diff before approving.
        title_changes = (
            db.query(ActivityLog)
              .join(SavedPoster, ActivityLog.target_id == SavedPoster.id, isouter=True)
              .filter(
                  ActivityLog.target_type == "saved_poster",
                  ActivityLog.action.in_(("saved", "deleted", "replaced")),
                  SavedPoster.master_title_id == t.id,
              )
              .order_by(ActivityLog.created_at.desc())
              .limit(50)
              .all()
        )
        pending_complete_blocks.append({
            "title": t,
            "revisions": revs_for_title,
            "changes": title_changes,
        })
        pending_title_ids.add(t.id)

    # ── Awaiting approval (standalone — NOT inside a pending completion) ────
    awaiting_q = scope_titles((
        db.query(Revision, SavedPoster, MasterTitle)
          .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
          .join(MasterTitle, SavedPoster.master_title_id == MasterTitle.id)
          .filter(Revision.status == "awaiting_approval")
          .order_by(Revision.submitted_at.desc().nullslast())
    ), proj)
    if pending_title_ids:
        # Hide any awaiting revision whose title is in a pending-completion
        # block — those are already rendered there.
        awaiting_q = awaiting_q.filter(
            ~MasterTitle.id.in_(pending_title_ids)
        )
    awaiting_rows = awaiting_q.all()

    # ── Open (waiting on user) ──────────────────────────────────────────────
    open_rows = scope_titles((
        db.query(Revision, SavedPoster, MasterTitle)
          .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
          .join(MasterTitle, SavedPoster.master_title_id == MasterTitle.id)
          .filter(Revision.status == "open")
    ), proj).order_by(Revision.created_at.desc()).all()

    # ── Recent mistake-deletions (legacy round-9 round-trip path) ────────────
    # After round 11 these only catch deletions on NON-flagged posters
    # (worker accidentally saved wrong image and deleted it). The
    # auto-resolve admin_verdict pattern is preserved for these.
    deletion_rows = scope_titles((
        db.query(Revision, SavedPoster, MasterTitle)
          .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
          .join(MasterTitle, SavedPoster.master_title_id == MasterTitle.id)
          .filter(
              Revision.status == "resolved",
              Revision.admin_acked_at.is_(None),
              Revision.admin_verdict.like("auto-resolved: file deleted%"),
          )
    ), proj).order_by(Revision.resolved_at.desc()).limit(50).all()

    # ── Resolved history ────────────────────────────────────────────────────
    resolved_rows = scope_titles((
        db.query(Revision, SavedPoster, MasterTitle)
          .join(SavedPoster, Revision.saved_poster_id == SavedPoster.id)
          .join(MasterTitle, SavedPoster.master_title_id == MasterTitle.id)
          .filter(Revision.status == "resolved")
    ), proj).order_by(Revision.resolved_at.desc()).limit(50).all()

    return templates.TemplateResponse(
        request,
        "admin_revisions.html",
        {"user": admin, "admin": admin,
            "pending_complete_blocks": pending_complete_blocks,
            "open_rows": open_rows, "awaiting_rows": awaiting_rows,
            "deletion_rows": deletion_rows,
            "resolved_rows": resolved_rows,
            "active_tab": "revisions",
        },
    )


# ── Skipped-title review ────────────────────────────────────────────────────

@router.get("/skipped", response_class=HTMLResponse)
def skipped_page(
    request: Request,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    rows = (
        scope_titles(db.query(MasterTitle), current_project(request, admin, db))
          .filter(MasterTitle.status == "skipped")
          .order_by(MasterTitle.completed_at.desc().nullslast(), MasterTitle.updated_at.desc())
          .limit(500)
          .all()
    )
    return templates.TemplateResponse(
        request,
        "admin_skipped.html",
        {"user": admin, "admin": admin, "titles": rows, "active_tab": "skipped"},
    )


# ── Deletion review (worker deleted a flagged poster) ───────────────────────

@router.post("/deletions/{revision_id}/acknowledge")
def acknowledge_deletion(
    revision_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Mark this auto-resolved-via-deletion revision as 'seen' by admin.
    It disappears from the Recent Deletions panel.
    """
    rev = db.query(Revision).filter_by(id=revision_id).first()
    if not rev:
        raise HTTPException(404, "Revision not found.")
    if rev.status != "resolved":
        raise HTTPException(400, "Revision is not resolved.")
    rev.admin_acked_at = datetime.utcnow()
    log_activity(db, user=admin, action="ack_deletion", target_type="revision", target_id=rev.id)
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/deletions/{revision_id}/escalate")
def escalate_deletion(
    revision_id: int,
    note: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin saw the deletion and disagrees. Sends the title back to the worker
    with a note (worker sees it as the admin-note banner on the active title).
    If the title was already 'complete', it's reverted to 'in_progress' so the
    worker actually sees it as something requiring action.
    """
    note = (note or "").strip()
    if not note:
        raise HTTPException(400, "A note is required when sending a deletion back.")
    rev = db.query(Revision).filter_by(id=revision_id).first()
    if not rev:
        raise HTTPException(404, "Revision not found.")
    sp = db.query(SavedPoster).filter_by(id=rev.saved_poster_id).first()
    if not sp:
        raise HTTPException(404, "Underlying poster row missing.")
    mt = db.query(MasterTitle).filter_by(id=sp.master_title_id).first()
    if not mt:
        raise HTTPException(404, "Underlying title row missing.")
    mt.admin_note = note
    # Re-flag the master so it shows up in master-list views and counters.
    # The escalation is an open issue from admin's POV until the worker
    # responds (re-saves something or admin manually clears).
    mt.needs_revision = 1
    if mt.status == "complete":
        mt.status = "in_progress"
        mt.completed_at = None
    rev.admin_acked_at = datetime.utcnow()
    log_activity(
        db, user=admin, action="escalate_deletion", target_type="master_title", target_id=mt.id,
        details={"revision_id": rev.id, "deleted_filename": sp.filename, "note": note},
    )
    db.commit()
    return JSONResponse({"ok": True})


# ── Backups + restore ───────────────────────────────────────────────────────

@router.get("/backups", response_class=HTMLResponse)
def backups_page(
    request: Request,
    admin: User = Depends(require_admin),
):
    from ..backups import list_backups
    return templates.TemplateResponse(
        request,
        "admin_backups.html",
        {"user": admin, "admin": admin, "backups": list_backups(),
         "active_tab": "backups"},
    )


@router.post("/backups/snapshot")
def create_snapshot(
    name: str = Form(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from ..backups import manual_snapshot
    try:
        path = manual_snapshot(name)
    except Exception as e:
        raise HTTPException(500, f"Snapshot failed: {e}")
    log_activity(db, user=admin, action="snapshot_created", target_type="backup",
                 details={"filename": path.name, "name": name})
    db.commit()
    return JSONResponse({"ok": True, "filename": path.name})


@router.post("/backups/restore")
def restore_from_backup(
    filename: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Replace the live DB with a backup file. Disposes the SQLAlchemy engine
    pool so the next request opens against the restored file.
    """
    from ..backups import restore_backup
    from ..db import engine
    # Log first while the old DB is still live.
    log_activity(db, user=admin, action="restore_started", target_type="backup",
                 details={"filename": filename})
    db.commit()
    db.close()

    # Dispose the engine pool so SQLite handles release the file.
    engine.dispose()

    try:
        safety = restore_backup(filename)
    except FileNotFoundError:
        raise HTTPException(404, "Backup not found.")
    except PermissionError:
        raise HTTPException(403, "Bad backup path.")
    except Exception as e:
        raise HTTPException(500, f"Restore failed: {e}")

    return JSONResponse({"ok": True, "safety_snapshot": safety.name if safety else None})


@router.post("/backups/delete")
def delete_backup_route(
    filename: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from ..backups import delete_backup
    try:
        ok = delete_backup(filename)
    except PermissionError:
        raise HTTPException(403, "Bad backup path.")
    if not ok:
        raise HTTPException(404, "Backup not found.")
    log_activity(db, user=admin, action="backup_deleted", target_type="backup",
                 details={"filename": filename})
    db.commit()
    return JSONResponse({"ok": True})


# ── Audit log ────────────────────────────────────────────────────────────────

@router.get("/audit", response_class=HTMLResponse)
def audit_page(request: Request, admin: User = Depends(require_admin)):
    return templates.TemplateResponse(
        request,
        "admin_audit.html",
        {"user": admin, "admin": admin, "active_tab": "audit"},
    )


@router.get("/api/audit")
def api_audit(
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=10, le=500),
    user: str = Query(""),
    action: str = Query(""),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    q = db.query(ActivityLog)
    if user.strip():
        q = q.filter(ActivityLog.username == user.strip())
    if action.strip():
        q = q.filter(ActivityLog.action == action.strip())
    total = q.count()
    rows = (
        q.order_by(ActivityLog.created_at.desc())
         .offset((page - 1) * page_size)
         .limit(page_size)
         .all()
    )
    return JSONResponse({
        "page": page, "page_size": page_size, "total": total,
        "pages": (total + page_size - 1) // page_size,
        "items": [
            {
                "id": r.id,
                "username": r.username, "action": r.action,
                "target_type": r.target_type, "target_id": r.target_id,
                "details": (json.loads(r.details) if r.details else None),
                "created_at": fmt_local(r.created_at, "%Y-%m-%d %H:%M:%S"),
            }
            for r in rows
        ],
    })


# ── Filesystem tree ─────────────────────────────────────────────────────────

@router.get("/api/tree")
def filesystem_tree(admin: User = Depends(require_admin)):
    out = []
    for username in list_users_with_workspaces():
        user_node = {"name": username, "children": []}
        for d in list_date_folders(username):
            d_path = WORKSPACE_DIR / username / d
            d_node = {"name": d, "children": []}
            for tf in sorted([p for p in d_path.iterdir() if p.is_dir()], key=lambda p: p.name):
                from ..utils import list_images_in
                imgs = list_images_in(tf)
                d_node["children"].append({"name": tf.name, "count": len(imgs)})
            user_node["children"].append(d_node)
        out.append(user_node)
    return JSONResponse({"workers": out})


# ── Zip a date folder ───────────────────────────────────────────────────────

ZIP_JOBS: dict[str, dict] = {}
ZIP_LOCK = threading.Lock()


def _zip_worker(job_id: str, sources: list[tuple[str, Path]], zip_path: Path):
    """
    Build a zip from one or more (date_label, src_dir) pairs.
    Files are written into per-date subfolders: `2026-05-15/title/file.jpg`.
    """
    try:
        all_files: list[tuple[Path, str]] = []  # (abs_path, arcname)
        for date_label, src in sources:
            for root, _dirs, files in os.walk(src):
                for f in files:
                    fp = Path(root) / f
                    # arcname = date/title/file.jpg
                    rel = fp.relative_to(src)
                    arcname = f"{date_label}/{rel}"
                    all_files.append((fp, arcname))
        total = max(len(all_files), 1)
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, (fp, arcname) in enumerate(all_files):
                zf.write(fp, arcname)
                with ZIP_LOCK:
                    ZIP_JOBS[job_id].update(done=i + 1, total=total)
        with ZIP_LOCK:
            ZIP_JOBS[job_id].update(state="done", path=str(zip_path), error="")
    except Exception as e:
        with ZIP_LOCK:
            ZIP_JOBS[job_id].update(state="error", error=str(e))


@router.post("/zip/start")
def zip_start(
    worker: str = Form(...),
    dates: str = Form(""),
    date: str = Form(""),
    admin: User = Depends(require_admin),
):
    """
    Start a background zip job for one or more dates.

    Accepts `dates` as a comma-separated string (e.g. "2026-05-15,2026-05-16").
    Also accepts legacy `date` (single value) for backwards compatibility.
    Files are written into per-date subfolders inside the zip.
    """
    # Merge both params: JS now sends `dates`, but keep `date` for any old callers.
    raw = dates or date or ""
    date_list = [d.strip() for d in raw.split(",") if d.strip()]
    if not date_list:
        raise HTTPException(400, "No dates specified.")

    sources: list[tuple[str, Path]] = []
    for d in date_list:
        src = (WORKSPACE_DIR / worker / d).resolve()
        if not safe_under_workspace(src) or not src.is_dir():
            continue  # skip missing folders silently
        sources.append((d, src))
    if not sources:
        raise HTTPException(404, "No matching date folders found on disk.")

    zip_dir = WORKSPACE_DIR / "_zips"
    zip_dir.mkdir(exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    if len(sources) == 1:
        zip_name = f"{worker}_{date_list[0]}_{ts}.zip"
    else:
        zip_name = f"{worker}_{date_list[0]}_to_{date_list[-1]}_{ts}.zip"
    zip_path = zip_dir / zip_name

    job_id = f"{worker}-zip-{ts}"
    with ZIP_LOCK:
        ZIP_JOBS[job_id] = {
            "state": "running", "done": 0, "total": 0,
            "path": "", "error": "", "name": zip_name,
        }
    threading.Thread(target=_zip_worker, args=(job_id, sources, zip_path), daemon=True).start()
    return JSONResponse({"ok": True, "job_id": job_id, "name": zip_name})


@router.get("/zip/status/{job_id}")
def zip_status(job_id: str, admin: User = Depends(require_admin)):
    with ZIP_LOCK:
        job = ZIP_JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Unknown job.")
    return JSONResponse(job)


@router.get("/zip/download/{job_id}")
def zip_download(job_id: str, admin: User = Depends(require_admin)):
    with ZIP_LOCK:
        job = ZIP_JOBS.get(job_id)
    if not job or job.get("state") != "done":
        raise HTTPException(404, "Zip not ready.")
    p = Path(job["path"])
    if not p.is_file():
        raise HTTPException(404, "Zip file missing.")
    return FileResponse(p, filename=p.name, media_type="application/zip")


# ═══════════════════════════════════════════════════════════════════════════
# Payments
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/payments", response_class=HTMLResponse)
def payments_page(
    request: Request,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin payments page. Shows:
      - Settings (rate KES, week start day) — editable.
      - Worker-by-worker counters for today and current week.
      - History of past PaymentRuns, with per-run "push to worker" toggle.
    """
    from ..payments import get_rate_kes, get_week_start_day, all_runs

    workers = (
        db.query(User)
          .filter(User.role == "worker", User.is_active == 1, User.is_deleted == 0)
          .order_by(User.username.asc())
          .all()
    )
    rate = get_rate_kes(db)
    week_start = get_week_start_day(db)
    runs = all_runs(db, limit=200)

    return templates.TemplateResponse(
        request, "admin_payments.html",
        {
            "user": admin, "admin": admin,
            "active_tab": "payments",
            "workers": workers,
            "rate_kes": rate,
            "week_start_day": week_start,
            "runs": runs,
            "today": local_today(),
        },
    )


@router.post("/payments/settings")
def payments_settings(
    rate_kes: str = Form(...),
    week_start_day: int = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Save the per-poster rate (KES) + week start day (0=Mon..6=Sun)."""
    from ..payments import set_setting, parse_decimal
    try:
        rate_dec = parse_decimal(rate_kes)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if rate_dec < 0:
        raise HTTPException(400, "Rate must be non-negative.")
    if not (0 <= week_start_day <= 6):
        raise HTTPException(400, "week_start_day must be 0..6.")
    set_setting(db, "pay_rate_kes",   str(rate_dec), by=admin.username)
    set_setting(db, "week_start_day", str(week_start_day), by=admin.username)
    log_activity(db, user=admin, action="payments_settings", target_type="settings",
                 details={"rate_kes": str(rate_dec), "week_start_day": week_start_day})
    db.commit()
    return RedirectResponse("/admin/payments", status_code=302)


@router.get("/api/payments/preview")
def payments_preview(
    worker_id: int = Query(...),
    start: str = Query(...),     # YYYY-MM-DD
    end:   str = Query(...),     # YYYY-MM-DD
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Preview eligible posters and total for a worker over [start, end].
    Used by the JS to compute the totals as admin picks dates.

    Also returns `unpaid_dates_before` — a list of {"date", "count"} for
    PAST days where this worker has eligible-but-unpaid posters. The UI
    surfaces these as a "you forgot these days" indicator so admin doesn't
    accidentally skip past-day backlog (P1).
    """
    from ..payments import (
        eligible_poster_ids, get_rate_kes, parse_decimal,
        unpaid_dates_before, local_today,
    )
    try:
        start_d = date.fromisoformat(start)
        end_d   = date.fromisoformat(end)
    except ValueError:
        raise HTTPException(400, "Bad date.")
    if end_d < start_d:
        raise HTTPException(400, "End is before start.")
    ids = eligible_poster_ids(db, worker_id=worker_id, start=start_d, end=end_d)
    rate = get_rate_kes(db)
    rate_dec = parse_decimal(rate)
    total = rate_dec * len(ids)
    # Per-day breakdown (saved-on dates) so admin can see distribution.
    by_day: dict[str, int] = {}
    if ids:
        rows = db.query(SavedPoster.id, SavedPoster.original_save_date).filter(SavedPoster.id.in_(ids)).all()
        for _id, d in rows:
            by_day[d.isoformat()] = by_day.get(d.isoformat(), 0) + 1

    # Unpaid-but-eligible from past days that aren't in the current preview range.
    today = local_today()
    unpaid_before = unpaid_dates_before(db, worker_id=worker_id, today=today)
    # Filter out days already inside [start_d, end_d] — those are visible
    # in `by_day` already; we want days OUTSIDE the picker's current view.
    unpaid_outside = [
        u for u in unpaid_before
        if not (start_d.isoformat() <= u["date"] <= end_d.isoformat())
    ]

    return JSONResponse({
        "ok": True,
        "worker_id":    worker_id,
        "start":        start_d.isoformat(),
        "end":          end_d.isoformat(),
        "poster_count": len(ids),
        "rate_kes":     str(rate_dec),
        "computed_total_kes": str(total),
        "by_day":       by_day,
        "poster_ids":   ids,
        "unpaid_before_outside_range": unpaid_outside,
    })


@router.post("/payments/mark_paid")
def payments_mark_paid(
    worker_id: int = Form(...),
    start: str = Form(...),
    end:   str = Form(...),
    amount_kes: str = Form(...),
    reference: str = Form(""),
    note: str = Form(""),
    push_to_worker: int = Form(0),
    include_back_pay_dates: str = Form(""),  # comma-separated YYYY-MM-DD list
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Write a PaymentRun for [start, end] for `worker_id`. Snapshots eligible
    poster IDs at this moment so future runs don't double-count them.
    Optionally pushes a receipt for worker acknowledgement.

    `include_back_pay_dates` lets admin pull in older days that have
    eligible-but-unpaid posters (e.g. "I paid Mon–Sun last week before
    realising Tuesday had a flagged poster that resolved later — include
    that Tuesday's posters in this run").
    """
    from ..payments import (
        eligible_poster_ids, get_rate_kes, parse_decimal, per_day_breakdown,
    )
    try:
        start_d = date.fromisoformat(start)
        end_d   = date.fromisoformat(end)
        amount  = parse_decimal(amount_kes)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if end_d < start_d:
        raise HTTPException(400, "End is before start.")
    if amount < 0:
        raise HTTPException(400, "Amount must be non-negative.")

    worker = db.query(User).filter_by(id=worker_id, role="worker").first()
    if worker is None:
        raise HTTPException(404, "Worker not found.")
    worker_username = worker.username

    # Primary eligible IDs from the chosen [start, end] range.
    ids = eligible_poster_ids(db, worker_id=worker_id, start=start_d, end=end_d)

    # Optionally fold in older "back-pay" days admin explicitly included.
    back_pay_dates: list[str] = []
    if include_back_pay_dates.strip():
        for raw in include_back_pay_dates.split(","):
            raw = raw.strip()
            if not raw:
                continue
            try:
                bp_d = date.fromisoformat(raw)
            except ValueError:
                raise HTTPException(400, f"Bad back-pay date: {raw}")
            # Each back-pay day is a separate one-day eligibility query.
            bp_ids = eligible_poster_ids(db, worker_id=worker_id, start=bp_d, end=bp_d)
            ids.extend(bp_ids)
            if bp_ids:
                back_pay_dates.append(bp_d.isoformat())

    # Dedup defensively (shouldn't have duplicates but safer).
    ids = sorted(set(ids))
    rate = get_rate_kes(db)

    # Per-day breakdown for receipt transparency.
    by_day = per_day_breakdown(db, poster_ids=ids)

    # ── Per-project pricing ─────────────────────────────────────────────────
    # ONE payment per worker — splitting the payout would mean two M-Pesa
    # transfers for the same week's work. What differs per project is the
    # RATE, so the amount is sum(count_in_project x rate_of_project) and the
    # receipt shows the working.
    from ..payments import price_run
    pricing = price_run(db, ids)
    by_project = pricing["by_project"]

    # The admin types the amount they actually sent, which is authoritative —
    # they may round, or add a bonus. But when the field is left at the
    # computed figure we record the per-project working alongside it so the
    # receipt can show HOW that number was reached across two rates.

    run = PaymentRun(
        worker_id           = worker_id,
        worker_username     = worker_username,
        period_start        = start_d,
        period_end          = end_d,
        poster_count        = len(ids),
        rate_kes            = str(parse_decimal(rate)),
        amount_kes          = str(amount),
        reference           = reference.strip() or None,
        note                = note.strip() or None,
        poster_ids_json     = json.dumps(ids),
        by_day_json         = json.dumps(by_day),
        by_project_json     = json.dumps(by_project) if by_project else None,
        back_pay_dates_json = json.dumps(back_pay_dates) if back_pay_dates else None,
        pushed_at           = datetime.utcnow() if push_to_worker else None,
        created_by          = admin.username,
    )
    db.add(run)
    db.flush()
    log_activity(
        db, user=admin, action="paid", target_type="payment_run", target_id=run.id,
        details={"worker": worker_username, "start": start, "end": end,
                 "count": len(ids), "amount": str(amount), "pushed": bool(push_to_worker),
                 "back_pay_dates": back_pay_dates},
    )

    # ── Pipeline hook: auto-greenlight what was just paid for ───────────
    # Paying for a batch is the natural signal that it's approved for
    # post-production, which is the workflow you already follow manually.
    # Driven off the run's poster_ids so back-pay days are included and
    # unpaid days never leak in. Respects the `greenlight_mode` setting, so
    # setting it to 'manual' disables this without touching code.
    #
    # Deliberately non-fatal: recording the payment must never fail because
    # of a pipeline problem (e.g. migration not yet run).
    greenlit = None
    try:
        from .. import pipeline as P
        greenlit = P.greenlight_for_payment_run(db, run, by=admin.username)
        if greenlit.get("greenlit"):
            log_activity(
                db, user=admin, action="pipeline_greenlight",
                target_type="payment_run", target_id=run.id,
                details={"via": "payment", **greenlit},
            )
    except Exception as e:
        greenlit = {"error": str(e)}

    db.commit()
    return JSONResponse({"ok": True, "run_id": run.id, "poster_count": len(ids),
                         "back_pay_dates": back_pay_dates,
                         "greenlit": greenlit})


@router.post("/payments/{run_id}/push")
def payments_push_run(
    run_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Push (or re-push) a receipt to the worker. Resets ack_at."""
    run = db.query(PaymentRun).filter_by(id=run_id).first()
    if run is None:
        raise HTTPException(404, "Run not found.")
    run.pushed_at = datetime.utcnow()
    run.ack_at = None           # clear any prior ack so worker has to confirm again
    run.not_received_at = None  # v15: also clear dispute flag on re-push
    log_activity(db, user=admin, action="receipt_push", target_type="payment_run", target_id=run.id)
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/payments/{run_id}/delete")
def payments_delete_run(
    run_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Delete a PaymentRun. Use only to fix mistakes — its poster IDs become
    eligible again. Gives admin a way to undo a wrong "MARK PAID" click.
    """
    run = db.query(PaymentRun).filter_by(id=run_id).first()
    if run is None:
        raise HTTPException(404, "Run not found.")
    snap = {"worker": run.worker_username, "amount": run.amount_kes,
            "count": run.poster_count, "period": f"{run.period_start}..{run.period_end}"}
    db.delete(run)
    log_activity(db, user=admin, action="payment_run_deleted",
                 target_type="payment_run", target_id=run_id, details=snap)
    db.commit()
    return JSONResponse({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════
# Chat (admin side)
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/chat", response_class=HTMLResponse)
def chat_page(
    request: Request,
    worker_id: int = Query(0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin chat hub: list of worker threads + an open thread on the right.

    We DON'T auto-select the first thread anymore. Auto-selecting + the
    chat.js auto-mark-read combo meant landing on this page from a badge
    notification immediately cleared the unread state before admin had a
    chance to actually read the message. Now admin must explicitly click
    a thread to open it.
    """
    from ..chat import admin_thread_summaries
    threads = admin_thread_summaries(db, viewer_id=admin.id)
    return templates.TemplateResponse(
        request, "admin_chat.html",
        {"user": admin, "admin": admin, "active_tab": "chat",
         "threads": threads, "selected_worker_id": worker_id},
    )


@router.get("/api/chat/{worker_id}")
def chat_thread(
    worker_id: int,
    after: int = Query(0, ge=0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from ..chat import list_messages, serialize_message
    rows = list_messages(db, worker_id=worker_id, after_id=(after or None), limit=200)
    return JSONResponse({"ok": True, "messages": [serialize_message(m) for m in rows]})


@router.post("/api/chat/{worker_id}/send")
def chat_admin_send(
    worker_id: int,
    body: str = Form(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from ..chat import send_message, serialize_message
    # Validate recipient exists.
    if not db.query(User).filter_by(id=worker_id, role="worker").first():
        raise HTTPException(404, "Worker not found.")
    try:
        msg = send_message(db, worker_id=worker_id, sender=admin, body=body)
    except ValueError as e:
        raise HTTPException(400, str(e))
    log_activity(db, user=admin, action="chat_sent", target_type="chat", target_id=msg.id,
                 details={"to_worker_id": worker_id})
    db.commit()
    return JSONResponse({"ok": True, "message": serialize_message(msg)})


@router.post("/api/chat/{worker_id}/mark_read")
def chat_admin_mark_read(
    worker_id: int,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from ..chat import mark_read
    mark_read(db, worker_id=worker_id, viewer_id=admin.id)
    db.commit()
    return JSONResponse({"ok": True})


@router.get("/api/chat/_summary")
def chat_summary(
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Polled by admin pages to render an unread badge in the topbar."""
    from ..chat import admin_thread_summaries
    threads = admin_thread_summaries(db, viewer_id=admin.id)
    total_unread = sum(t["unread"] for t in threads)
    return JSONResponse({"ok": True, "total_unread": total_unread, "threads": threads})


# ═══════════════════════════════════════════════════════════════════════════
# Activity log v2 — readable, with optional comments-only filter
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/api/activity")
def api_activity(
    since_id: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=500),
    user_filter: str = Query(""),
    action_filter: str = Query(""),
    comments_only: int = Query(0, ge=0, le=1),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    JSON activity log with optional filters. The page uses this to render
    the new readable list and to power the "Comments only" view.
    """
    q = db.query(ActivityLog).order_by(ActivityLog.id.desc())
    if since_id:
        q = q.filter(ActivityLog.id < since_id)  # paginate older
    if user_filter:
        q = q.filter(ActivityLog.username == user_filter)
    if action_filter:
        q = q.filter(ActivityLog.action == action_filter)
    rows = q.limit(limit).all()

    out = []
    # Batch-fetch title names for master_title targets so the activity log
    # can show "Dogma (1999)" instead of "master#1190".
    mt_ids = {r.target_id for r in rows if r.target_type == "master_title" and r.target_id}
    mt_names = {}
    if mt_ids:
        for mt in db.query(MasterTitle).filter(MasterTitle.id.in_(mt_ids)).all():
            mt_names[mt.id] = f"{mt.title} ({mt.year})" if mt.year else mt.title

    for r in rows:
        comment = ""
        details = {}
        if r.details:
            try:
                details = json.loads(r.details)
            except (TypeError, ValueError):
                details = {}
        for key in ("comment", "reason", "note", "verdict", "worker_note", "admin_note"):
            v = details.get(key)
            if v:
                comment = str(v)
                break
        if comments_only and not comment:
            continue
        # Inject the title name for master_title targets if available.
        if r.target_type == "master_title" and r.target_id in mt_names:
            details["title"] = mt_names[r.target_id]
        out.append({
            "id":          r.id,
            "created_at":  fmt_local(r.created_at, "%Y-%m-%d %H:%M:%S"),
            "username":    r.username or "",
            "action":      r.action,
            "target_type": r.target_type or "",
            "target_id":   r.target_id,
            "comment":     comment,
            "details":     details,
        })
    return JSONResponse({"ok": True, "rows": out})


# ── Admin peek mode (read-only view of worker's dashboard) ─────────────────

@router.get("/peek", response_class=HTMLResponse)
def peek_list(
    request: Request,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """List of workers the admin can peek at."""
    workers = (
        db.query(User)
          .filter(User.role == "worker", User.is_deleted == 0)
          .order_by(User.username.asc())
          .all()
    )
    return templates.TemplateResponse(
        request,
        "admin_peek_list.html",
        {"user": admin, "admin": admin, "workers": workers, "active_tab": "peek"},
    )


@router.get("/peek/{username}", response_class=HTMLResponse)
def peek_worker(
    request: Request,
    username: str,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Render the actual worker dashboard template (user_dashboard.html) in
    read-only peek mode. The admin sees exactly what the worker sees —
    same layout, same poster sizes, same flag cards — but all buttons and
    inputs are disabled via a CSS overlay. The JS fetches state from the
    admin peek API instead of /api/state.
    """
    worker = db.query(User).filter_by(username=username, role="worker").first()
    if not worker:
        raise HTTPException(404, "Worker not found.")
    # Peek is a project-level page, so it shows the worker's state WITHIN the
    # project the admin is currently in — not a merged view of every project
    # that worker touches.
    from .worker import _state_payload
    state = _state_payload(db, worker, current_project(request, admin, db))
    return templates.TemplateResponse(
        request,
        "user_dashboard.html",
        {"user": admin, "admin": admin, "state": state,
         "peek_username": worker.username, "active_tab": "peek"},
    )


@router.get("/api/peek/{username}")
def api_peek_worker(
    request: Request,
    username: str,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """JSON state for a worker — same shape as /api/state but fetched by admin."""
    worker = db.query(User).filter_by(username=username, role="worker").first()
    if not worker:
        raise HTTPException(404, "Worker not found.")
    from .worker import _state_payload
    return JSONResponse(_state_payload(db, worker, current_project(request, admin, db)))


# ── Per-worker stats panel ──────────────────────────────────────────────────

@router.get("/stats", response_class=HTMLResponse)
def admin_stats_page(
    request: Request,
    worker_id: int = Query(0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Admin's view of worker performance. Sidebar lists workers; clicking
    one shows their full stats panel with the same chart + records the
    worker sees, plus admin-only flag rate + turnaround.
    """
    workers = (
        db.query(User)
          .filter(User.role == "worker", User.is_deleted == 0)
          .order_by(User.username.asc())
          .all()
    )
    if not worker_id and workers:
        worker_id = workers[0].id
    return templates.TemplateResponse(
        request, "admin_stats.html",
        {"user": admin, "admin": admin,
         "active_tab": "stats",
         "stats_scope": "project",
         "workers": workers,
         # Offered as a FILTER, not applied. Even standing inside a project,
         # "how is this worker doing" is usually the whole-person question.
         "projects": db.query(Project).order_by(Project.id.asc()).all(),
         "selected_worker_id": worker_id},
    )


@router.get("/master_stats", response_class=HTMLResponse)
def admin_master_stats_page(
    request: Request,
    worker_id: int = Query(0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    The same worker-performance page, reached from the master nav.

    Worker output is a person-level fact, not a project-level one — a worker
    covering two niches has one throughput and one flag rate. So rather than
    build a second stats page, this renders the same one under the master nav
    so the admin doesn't have to enter an arbitrary project to answer "how is
    humphrey doing".
    """
    workers = (
        db.query(User)
          .filter(User.role == "worker", User.is_deleted == 0)
          .order_by(User.username.asc())
          .all()
    )
    if not worker_id and workers:
        worker_id = workers[0].id
    return templates.TemplateResponse(
        request, "admin_stats.html",
        {"user": admin, "admin": admin,
         "active_tab": "master_stats",
         "stats_scope": "master",
         "workers": workers,
         "projects": db.query(Project).order_by(Project.id.asc()).all(),
         "selected_worker_id": worker_id},
    )


@router.get("/api/stats/{worker_id}")
def api_admin_stats(
    worker_id: int,
    project_id: int = Query(0),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """JSON stats for one worker — admin view (includes admin_only block)."""
    from ..stats import compute_worker_stats
    # No project_id means every project, which is what the page asks for by
    # default: a worker's throughput is one number regardless of how many
    # niches they cover. The filter narrows it when you want to compare.
    data = compute_worker_stats(db, worker_id=worker_id, is_admin_view=True,
                                project_id=project_id or None)
    return JSONResponse(data)
